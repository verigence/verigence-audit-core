from io import BytesIO

from openpyxl import Workbook

from audit_core.mahindra_native_workbooks import (
    parse_native_discount_policy_workbook,
    parse_native_segment_workbook,
)


def _bytes(workbook: Workbook) -> bytes:
    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def test_parses_native_pv_price_sheet() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "ScorpioN Diesel"
    sheet.cell(3, 2, "Price list w.e.f. 04th Aug'26")
    sheet.cell(5, 5, "SCORPIO N")
    for col, value in {
        3: "TRIM",
        4: "FUEL",
        5: "TRANSMISSION",
        6: "DRIVE",
        7: "SEATER",
        9: "Ex-showroom Price",
        10: "Tax Collection at Source (TCS)",
        11: "Insurance 30%",
    }.items():
        sheet.cell(7, col, value)
    sheet.cell(10, 3, "Z8 L")
    sheet.cell(10, 4, "Diesel")
    sheet.cell(10, 5, "AT")
    sheet.cell(10, 6, "4WD")
    sheet.cell(10, 7, 7)
    sheet.cell(10, 9, 2425144.64)
    sheet.cell(10, 10, 24251.45)
    sheet.cell(10, 11, 120237.52)

    rows = parse_native_segment_workbook(
        _bytes(workbook),
        "PASSENGER_VEHICLE",
    )

    assert len(rows) == 3
    parsed = rows[0][1]
    assert parsed["model_name"] == "SCORPIO N"
    assert parsed["trim_name"] == "Z8 L"
    assert parsed["configuration_code"] == "SCORPIO_N_Z8_L_DIESEL_AT_4WD_7"
    assert parsed["component_key"] == "EX_SHOWROOM_PRICE"
    assert parsed["source_effective_from"] == "2026-08-04"


def test_parses_native_bev_price_sheet() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "BE6"
    sheet.cell(3, 1, "Price list w.e.f. 04.08.2026")
    sheet.cell(7, 1, "Model & Variant")
    sheet.cell(7, 5, "Ex-showroom Price")
    sheet.cell(7, 6, "Tax Collection at Source (TCS)")
    sheet.cell(8, 2, "STR")
    sheet.cell(10, 1, "BE 6 One B59 R18 NCH")
    sheet.cell(10, 2, 5)
    sheet.cell(10, 5, 1889999.99)
    sheet.cell(10, 6, 18899.99)

    rows = parse_native_segment_workbook(
        _bytes(workbook),
        "BATTERY_ELECTRIC",
    )

    assert len(rows) == 2
    parsed = rows[0][1]
    assert parsed["model_name"] == "BE6"
    assert parsed["fuel_powertrain"] == "ELECTRIC"
    assert parsed["seating_capacity"] == 5
    assert parsed["source_effective_from"] == "2026-08-04"


def test_parses_native_discount_grid_and_tradein_policy() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Grid for Aug"
    sheet.cell(1, 2, "Effective 1st Aug 2026")
    sheet.cell(2, 3, "Booking Protection")
    sheet.cell(2, 4, "Agreed Buffer")
    sheet.cell(2, 5, "Insurance OD %")
    sheet.cell(3, 2, "3XO")
    sheet.cell(3, 3, "30 days")
    sheet.cell(3, 4, 5000)
    sheet.cell(3, 5, 0.5)
    sheet.cell(28, 2, "Trade-in")
    sheet.cell(
        28,
        3,
        "Used vehicles accepted under Trade-in must be retailed within 90 days. "
        "Minimum profit to be maintained: ₹10,000 for Passenger Vehicles (PV) "
        "and ₹5,000 for Commercial Vehicles (CV).",
    )

    rows = parse_native_discount_policy_workbook(_bytes(workbook))
    values = {row[1]["parameter_key"]: row[1] for row in rows}

    assert values["BOOKING_PROTECTION_DAYS"]["value_number"] == "30"
    assert values["AGREED_BUFFER"]["value_number"] == "5000"
    assert values["INSURANCE_OD_PERCENT"]["value_number"] == "50.0"
    assert values["TRADE_IN_MAX_HOLDING_DAYS"]["value_number"] == "90"
    assert values["TRADE_IN_MIN_PROFIT_PV"]["value_number"] == "10000"
    assert values["TRADE_IN_MIN_PROFIT_CV"]["value_number"] == "5000"
