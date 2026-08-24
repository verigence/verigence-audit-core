from __future__ import annotations

import os
from io import BytesIO
from uuid import uuid4

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import create_engine, text

from audit_core.dependencies import HumanAdminRequest, require_project_admin_request
from audit_core.main import app
from audit_core.project_master_imports import build_template
from audit_core.security_integration import SecurityAdminContext


@pytest.fixture
def master_setup():
    database_url = os.environ.get("DATABASE_URL")
    if not database_url:
        pytest.skip("DATABASE_URL is required for UC02 Project Master integration tests")

    engine = create_engine(database_url)
    suffix = uuid4().hex
    tenant_id = f"tenant-master-api-{suffix}"
    with engine.begin() as connection:
        category_id = connection.execute(
            text(
                "INSERT INTO auditcore.product_categories (category_code, category_name) "
                "VALUES (:code, :name) RETURNING product_category_id"
            ),
            {"code": f"API-CAT-{suffix}", "name": "Vehicle"},
        ).scalar_one()
        oem_id = connection.execute(
            text(
                "INSERT INTO auditcore.oems (oem_code, oem_name) "
                "VALUES (:code, :name) RETURNING oem_id"
            ),
            {"code": f"API-OEM-{suffix}", "name": "Master API OEM"},
        ).scalar_one()
        connection.execute(
            text(
                """
                INSERT INTO auditcore.projects (
                    tenant_id, project_code, project_name, oem_id,
                    product_category_id, effective_start_date, project_status
                ) VALUES (
                    :tenant_id, :code, 'Master API Project', :oem_id,
                    :category_id, DATE '2026-08-01', 'CONFIGURING'
                )
                """
            ),
            {
                "tenant_id": tenant_id,
                "code": f"API-PROJ-{suffix}",
                "oem_id": oem_id,
                "category_id": category_id,
            },
        )

    admin_request = HumanAdminRequest(
        user_id="superadmin-master",
        bearer_token="master-human-token",
        admin_context=SecurityAdminContext(
            user_id="superadmin-master",
            is_super_admin=True,
            admin_scopes=(),
        ),
    )
    app.dependency_overrides[require_project_admin_request] = lambda: admin_request
    try:
        yield {
            "engine": engine,
            "tenant_id": tenant_id,
            "suffix": suffix,
        }
    finally:
        app.dependency_overrides.clear()
        engine.dispose()


def _workbook(master_key: str, row: dict[str, object]) -> bytes:
    template = build_template(master_key)
    workbook = load_workbook(BytesIO(template))
    data_sheet_name = next(name for name in workbook.sheetnames if name != "_meta")
    sheet = workbook[data_sheet_name]
    headers = [cell.value for cell in sheet[1]]
    sheet.append([row.get(str(header)) for header in headers])
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _upload(
    client: TestClient,
    *,
    tenant_id: str,
    master_key: str,
    content: bytes,
    idempotency_key: str,
    effective_from: str | None = "2026-08-01",
):
    data = {}
    if effective_from is not None:
        data["effectiveFrom"] = effective_from
    return client.post(
        f"/v1/tenants/{tenant_id}/project-masters/AUDIT_CORE/{master_key}/imports",
        headers={"Idempotency-Key": idempotency_key},
        data=data,
        files={
            "file": (
                f"{master_key.lower()}.xlsx",
                content,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )


def test_product_master_upload_preview_confirm_publish_and_catalogue(master_setup) -> None:
    setup = master_setup
    client = TestClient(app, raise_server_exceptions=False)
    sku_code = f"SKU-{setup['suffix']}"
    content = _workbook(
        "PRODUCT_MASTER",
        {
            "sku_code": sku_code,
            "model_code": f"MODEL-{setup['suffix']}",
            "model_name": "Model A",
            "model_year": 2026,
            "variant_code": f"VAR-{setup['suffix']}",
            "variant_name": "Variant A",
            "fuel_powertrain": "ICE",
            "transmission": "AT",
            "body_type": "SUV",
            "colour_code": f"CLR-{setup['suffix']}",
            "colour_name": "Blue",
            "sku_attributes_json": '{"trim":"top"}',
            "variant_attributes_json": '{"doors":5}',
        },
    )

    before = setup["engine"].connect().execute(
        text(
            "SELECT count(*) FROM auditcore.project_product_master_versions "
            "WHERE tenant_id=:tenant_id"
        ),
        {"tenant_id": setup["tenant_id"]},
    ).scalar_one()
    assert before == 0

    upload = _upload(
        client,
        tenant_id=setup["tenant_id"],
        master_key="PRODUCT_MASTER",
        content=content,
        idempotency_key="product-import-1",
    )
    assert upload.status_code == 201, upload.text
    summary = upload.json()
    assert summary["status"] == "PREVIEW_READY"
    assert summary["rowsParsed"] == 1
    assert summary["errorRows"] == 0

    replay = _upload(
        client,
        tenant_id=setup["tenant_id"],
        master_key="PRODUCT_MASTER",
        content=content,
        idempotency_key="product-import-1",
    )
    assert replay.status_code == 201
    assert replay.json()["importId"] == summary["importId"]

    rows = client.get(
        f"/v1/tenants/{setup['tenant_id']}/project-master-imports/{summary['importId']}/rows"
    )
    assert rows.status_code == 200
    assert rows.json()["items"][0]["parsedData"]["sku_code"] == sku_code

    with setup["engine"].connect() as connection:
        assert connection.execute(
            text(
                "SELECT count(*) FROM auditcore.project_product_master_versions "
                "WHERE tenant_id=:tenant_id"
            ),
            {"tenant_id": setup["tenant_id"]},
        ).scalar_one() == 0

    confirmed = client.post(
        f"/v1/tenants/{setup['tenant_id']}/project-master-imports/{summary['importId']}/confirm"
    )
    assert confirmed.status_code == 200, confirmed.text
    confirmed_body = confirmed.json()
    assert confirmed_body["status"] == "CONFIRMED"
    assert confirmed_body["confirmedVersionId"]

    versions = client.get(
        f"/v1/tenants/{setup['tenant_id']}/project-masters/AUDIT_CORE/PRODUCT_MASTER/versions"
    )
    assert versions.status_code == 200
    assert versions.json()[0]["lifecycleStatus"] == "DRAFT"

    published = client.post(
        f"/v1/tenants/{setup['tenant_id']}/project-masters/AUDIT_CORE/PRODUCT_MASTER/"
        f"versions/{confirmed_body['confirmedVersionId']}/publish"
    )
    assert published.status_code == 200, published.text
    assert published.json()["lifecycleStatus"] == "PUBLISHED"

    versions = client.get(
        f"/v1/tenants/{setup['tenant_id']}/project-masters/AUDIT_CORE/PRODUCT_MASTER/versions"
    )
    assert versions.status_code == 200
    assert versions.json()[0]["lifecycleStatus"] == "PUBLISHED"

    # The unqualified catalogue is the cross-module UC02 facade. It must not silently
    # hide the three DI-owned master domains when DI administration is unconfigured.
    catalogue = client.get(f"/v1/tenants/{setup['tenant_id']}/project-masters")
    assert catalogue.status_code == 503
    assert catalogue.json()["errorCode"] == "VAC-DI-001"


def test_price_list_requires_effective_product_master_context(master_setup) -> None:
    setup = master_setup
    client = TestClient(app, raise_server_exceptions=False)
    sku_code = f"PRICE-SKU-{setup['suffix']}"
    product_content = _workbook(
        "PRODUCT_MASTER",
        {
            "sku_code": sku_code,
            "model_code": f"PRICE-MODEL-{setup['suffix']}",
            "model_name": "Price Model",
            "variant_code": f"PRICE-VAR-{setup['suffix']}",
            "variant_name": "Price Variant",
        },
    )
    product_upload = _upload(
        client,
        tenant_id=setup["tenant_id"],
        master_key="PRODUCT_MASTER",
        content=product_content,
        idempotency_key="product-for-price",
    )
    product_confirm = client.post(
        f"/v1/tenants/{setup['tenant_id']}/project-master-imports/"
        f"{product_upload.json()['importId']}/confirm"
    )
    product_version_id = product_confirm.json()["confirmedVersionId"]

    price_content = _workbook(
        "PRICE_LIST",
        {
            "price_list_code": f"PL-{setup['suffix']}",
            "price_list_name": "Retail Price",
            "sku_code": sku_code,
            "component_key": "EX_SHOWROOM",
            "standard_amount": 1000000,
            "currency_code": "INR",
        },
    )
    before_publish = _upload(
        client,
        tenant_id=setup["tenant_id"],
        master_key="PRICE_LIST",
        content=price_content,
        idempotency_key="price-before-product-publish",
    )
    assert before_publish.status_code == 201
    assert before_publish.json()["status"] == "VALIDATION_FAILED"
    assert before_publish.json()["errorRows"] == 1

    product_publish = client.post(
        f"/v1/tenants/{setup['tenant_id']}/project-masters/AUDIT_CORE/PRODUCT_MASTER/"
        f"versions/{product_version_id}/publish"
    )
    assert product_publish.status_code == 200

    price_upload = _upload(
        client,
        tenant_id=setup["tenant_id"],
        master_key="PRICE_LIST",
        content=price_content,
        idempotency_key="price-after-product-publish",
    )
    assert price_upload.status_code == 201, price_upload.text
    assert price_upload.json()["status"] == "PREVIEW_READY"
    confirmed = client.post(
        f"/v1/tenants/{setup['tenant_id']}/project-master-imports/"
        f"{price_upload.json()['importId']}/confirm"
    )
    assert confirmed.status_code == 200, confirmed.text
    assert confirmed.json()["confirmedVersionId"]


def test_effective_dated_excel_master_requires_blank_by_default_wef(master_setup) -> None:
    setup = master_setup
    client = TestClient(app, raise_server_exceptions=False)
    content = _workbook(
        "PRODUCT_MASTER",
        {
            "sku_code": f"WEF-SKU-{setup['suffix']}",
            "model_code": f"WEF-MODEL-{setup['suffix']}",
            "model_name": "WEF Model",
            "variant_code": f"WEF-VAR-{setup['suffix']}",
            "variant_name": "WEF Variant",
        },
    )
    response = _upload(
        client,
        tenant_id=setup["tenant_id"],
        master_key="PRODUCT_MASTER",
        content=content,
        idempotency_key="missing-wef",
        effective_from=None,
    )
    assert response.status_code == 400
    assert response.json()["errorCode"] == "VAC-VAL-001"


def test_template_download_and_confirmed_import_cannot_be_deleted(master_setup) -> None:
    setup = master_setup
    client = TestClient(app, raise_server_exceptions=False)
    template = client.get(
        f"/v1/tenants/{setup['tenant_id']}/project-masters/AUDIT_CORE/PRODUCT_MASTER/template"
    )
    assert template.status_code == 200
    assert template.content.startswith(b"PK")

    content = _workbook(
        "PRODUCT_MASTER",
        {
            "sku_code": f"DELETE-SKU-{setup['suffix']}",
            "model_code": f"DELETE-MODEL-{setup['suffix']}",
            "model_name": "Delete Model",
            "variant_code": f"DELETE-VAR-{setup['suffix']}",
            "variant_name": "Delete Variant",
        },
    )
    uploaded = _upload(
        client,
        tenant_id=setup["tenant_id"],
        master_key="PRODUCT_MASTER",
        content=content,
        idempotency_key="confirmed-delete",
    )
    import_id = uploaded.json()["importId"]
    confirmed = client.post(
        f"/v1/tenants/{setup['tenant_id']}/project-master-imports/{import_id}/confirm"
    )
    assert confirmed.status_code == 200
    deleted = client.delete(
        f"/v1/tenants/{setup['tenant_id']}/project-master-imports/{import_id}"
    )
    assert deleted.status_code == 409