from __future__ import annotations

import re
from collections import Counter, defaultdict
from datetime import date
from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import Any

from openpyxl import load_workbook

_SEGMENT_MASTER_KEY = "MAHINDRA_SEGMENT_MASTER"
_DISCOUNT_POLICY_KEY = "DISCOUNT_POLICY"


class NativeWorkbookNotRecognized(Exception):
    pass


def _clean(value: Any) -> str:
    if value is None:
        return ""
    return " ".join(str(value).replace("\n", " ").split()).strip()


def _code(value: Any, limit: int = 80) -> str:
    text = re.sub(r"[^A-Z0-9]+", "_", _clean(value).upper()).strip("_")
    return (text or "UNKNOWN")[:limit]


def _number(value: Any) -> Decimal | None:
    if value is None or _clean(value) == "" or isinstance(value, bool):
        return None
    try:
        return Decimal(str(value).replace(",", "").replace("₹", "").strip())
    except (InvalidOperation, ValueError):
        return None


def _extract_wef(workbook) -> str | None:
    numeric_pattern = r"(\d{1,2})[.\-/](\d{1,2})[.\-/](\d{4})"
    text_pattern = r"(\d{1,2})(?:st|nd|rd|th)?\s+([A-Za-z]{3,9})['’ ]?(\d{2,4})"
    months = {
        "jan": 1,
        "january": 1,
        "feb": 2,
        "february": 2,
        "mar": 3,
        "march": 3,
        "apr": 4,
        "april": 4,
        "may": 5,
        "jun": 6,
        "june": 6,
        "jul": 7,
        "july": 7,
        "aug": 8,
        "august": 8,
        "sep": 9,
        "sept": 9,
        "september": 9,
        "oct": 10,
        "october": 10,
        "nov": 11,
        "november": 11,
        "dec": 12,
        "december": 12,
    }
    for sheet in workbook.worksheets[:3]:
        for row in sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 6), values_only=True):
            for value in row:
                text = _clean(value)
                if not text:
                    continue
                match = re.search(numeric_pattern, text)
                if match:
                    day, month, year = map(int, match.groups())
                    try:
                        return date(year, month, day).isoformat()
                    except ValueError:
                        pass
                match = re.search(text_pattern, text, re.IGNORECASE)
                if match:
                    day = int(match.group(1))
                    month = months.get(match.group(2).lower())
                    year = int(match.group(3))
                    if year < 100:
                        year += 2000
                    if month:
                        try:
                            return date(year, month, day).isoformat()
                        except ValueError:
                            pass
    return None


def _is_internal_template(workbook, master_key: str) -> bool:
    if "_meta" in workbook.sheetnames:
        return True
    if len(workbook.sheetnames) != 1:
        return False
    sheet = workbook[workbook.sheetnames[0]]
    headers = {_clean(cell.value) for cell in sheet[1] if _clean(cell.value)}
    if master_key == _SEGMENT_MASTER_KEY:
        return {
            "sku_code",
            "model_code",
            "trim_code",
            "configuration_code",
            "component_key",
            "standard_amount",
        }.issubset(headers)
    return {"scope_type", "parameter_key", "value_type"}.issubset(headers)


def _header_label(sheet, col: int) -> str:
    top = _clean(sheet.cell(7, col).value)
    leaf = _clean(sheet.cell(8, col).value)
    label = leaf or top
    if not label:
        return ""
    if re.fullmatch(r"\([A-Z ]+\)(?:=.*)?", label, re.IGNORECASE):
        label = top
    return label


def _component_columns(sheet, start_col: int) -> list[tuple[int, str, str]]:
    raw: list[tuple[int, str]] = []
    for col in range(start_col, sheet.max_column + 1):
        label = _header_label(sheet, col)
        if not label:
            continue
        has_numeric_value = any(
            _number(sheet.cell(row, col).value) is not None
            for row in range(10, min(sheet.max_row, 80) + 1)
        )
        if has_numeric_value:
            raw.append((col, label))

    counts = Counter(_code(label, 100) for _, label in raw)
    seen: defaultdict[str, int] = defaultdict(int)
    result: list[tuple[int, str, str]] = []
    for col, label in raw:
        base = _code(label, 92)
        seen[base] += 1
        key = base if counts[base] == 1 else f"{base}_{seen[base]}"
        result.append((col, key[:100], label))
    return result


def _data_rows_pv_cv(sheet) -> list[dict[str, Any]]:
    model_name = _clean(sheet.cell(5, 5).value) or _clean(sheet.title)
    model_code = _code(model_name, 60)
    component_columns = _component_columns(sheet, 9)
    if not component_columns:
        raise NativeWorkbookNotRecognized(f"No price columns found in sheet {sheet.title}")

    identities: list[tuple[int, dict[str, Any]]] = []
    for row_no in range(10, sheet.max_row + 1):
        trim = _clean(sheet.cell(row_no, 3).value)
        fuel = _clean(sheet.cell(row_no, 4).value)
        transmission = _clean(sheet.cell(row_no, 5).value)
        drive = _clean(sheet.cell(row_no, 6).value)
        seats = sheet.cell(row_no, 7).value
        if not trim:
            continue
        if not any(
            _number(sheet.cell(row_no, col).value) is not None
            for col, _, _ in component_columns
        ):
            continue
        identities.append(
            (
                row_no,
                {
                    "model_code": model_code,
                    "model_name": model_name,
                    "trim_code": _code(trim, 60),
                    "trim_name": trim,
                    "fuel_powertrain": fuel or None,
                    "transmission": transmission or None,
                    "drive_type": drive or None,
                    "seating_capacity": seats,
                    "body_type": None,
                },
            )
        )

    base_counts = Counter(
        (
            item["model_code"],
            item["trim_code"],
            _code(item.get("fuel_powertrain")),
            _code(item.get("transmission")),
            _code(item.get("drive_type")),
            _clean(item.get("seating_capacity")),
        )
        for _, item in identities
    )
    occurrences: defaultdict[tuple[Any, ...], int] = defaultdict(int)
    out: list[dict[str, Any]] = []
    for row_no, identity in identities:
        base_tuple = (
            identity["model_code"],
            identity["trim_code"],
            _code(identity.get("fuel_powertrain")),
            _code(identity.get("transmission")),
            _code(identity.get("drive_type")),
            _clean(identity.get("seating_capacity")),
        )
        occurrences[base_tuple] += 1
        suffix = f"_OPT{occurrences[base_tuple]}" if base_counts[base_tuple] > 1 else ""
        config_base = "_".join(
            part
            for part in [
                identity["model_code"],
                identity["trim_code"],
                _code(identity.get("fuel_powertrain"), 20),
                _code(identity.get("transmission"), 15),
                _code(identity.get("drive_type"), 15),
                _code(identity.get("seating_capacity"), 8),
            ]
            if part and part != "UNKNOWN"
        )
        configuration_code = (config_base + suffix)[:120]
        sku_code = configuration_code[:120]
        for col, component_key, component_label in component_columns:
            amount = _number(sheet.cell(row_no, col).value)
            if amount is None:
                continue
            row = dict(identity)
            row.update(
                {
                    "sku_code": sku_code,
                    "configuration_code": configuration_code,
                    "colour_code": None,
                    "colour_name": None,
                    "currency_code": "INR",
                    "component_key": component_key,
                    "component_label": component_label,
                    "standard_amount": str(amount),
                    "source_sheet": sheet.title,
                    "source_row": row_no,
                }
            )
            out.append(row)
    return out


def _data_rows_bev(sheet) -> list[dict[str, Any]]:
    model_name = _clean(sheet.title)
    model_code = _code(model_name, 60)
    component_columns = _component_columns(sheet, 5)
    if not component_columns:
        raise NativeWorkbookNotRecognized(f"No price columns found in sheet {sheet.title}")
    out: list[dict[str, Any]] = []
    for row_no in range(10, sheet.max_row + 1):
        vehicle = _clean(sheet.cell(row_no, 1).value)
        if not vehicle:
            continue
        if not any(
            _number(sheet.cell(row_no, col).value) is not None
            for col, _, _ in component_columns
        ):
            continue
        seats = sheet.cell(row_no, 2).value
        trim_name = vehicle
        trim_code = _code(vehicle, 60)
        configuration_code = _code(vehicle, 120)
        sku_code = configuration_code
        for col, component_key, component_label in component_columns:
            amount = _number(sheet.cell(row_no, col).value)
            if amount is None:
                continue
            out.append(
                {
                    "sku_code": sku_code,
                    "model_code": model_code,
                    "model_name": model_name,
                    "model_year": None,
                    "trim_code": trim_code,
                    "trim_name": trim_name,
                    "configuration_code": configuration_code,
                    "fuel_powertrain": "ELECTRIC",
                    "transmission": "AT",
                    "drive_type": None,
                    "seating_capacity": seats,
                    "body_type": None,
                    "colour_code": None,
                    "colour_name": None,
                    "currency_code": "INR",
                    "component_key": component_key,
                    "component_label": component_label,
                    "standard_amount": str(amount),
                    "source_sheet": sheet.title,
                    "source_row": row_no,
                }
            )
    return out


def parse_native_segment_workbook(content: bytes, segment_code: str | None = None):
    workbook = load_workbook(BytesIO(content), read_only=False, data_only=True)
    if _is_internal_template(workbook, _SEGMENT_MASTER_KEY):
        raise NativeWorkbookNotRecognized

    normalized: list[tuple[int, dict[str, Any]]] = []
    source_wef = _extract_wef(workbook)
    synthetic_row = 2
    recognized_sheets = 0
    for sheet in workbook.worksheets:
        bev_layout = (
            "MODEL" in _clean(sheet.cell(7, 1).value).upper()
            and "VARIANT" in _clean(sheet.cell(7, 1).value).upper()
        )
        pv_cv_layout = _clean(sheet.cell(7, 3).value).upper() == "TRIM"
        if not (bev_layout or pv_cv_layout):
            continue
        rows = _data_rows_bev(sheet) if bev_layout else _data_rows_pv_cv(sheet)
        if rows:
            recognized_sheets += 1
        for row in rows:
            row["source_effective_from"] = source_wef
            row["source_segment_code"] = segment_code
            normalized.append((synthetic_row, row))
            synthetic_row += 1
    if recognized_sheets == 0 or not normalized:
        raise NativeWorkbookNotRecognized
    return normalized


def _policy_number(text: str, pattern: str) -> Decimal | None:
    match = re.search(pattern, text, re.IGNORECASE)
    if not match:
        return None
    try:
        return Decimal(match.group(1).replace(",", ""))
    except InvalidOperation:
        return None


def _policy_segment(scope_key: str) -> str:
    scope_code = _code(scope_key, 80)
    if any(token in scope_code for token in ("BE6", "9E", "9S", "3XO_EV", "XUV400")):
        return "BATTERY_ELECTRIC"
    if any(token in scope_code for token in ("PICKUP", "MAXX", "SUPRO", "CAMPER", "VEERO")):
        return "COMMERCIAL"
    return "PASSENGER_VEHICLE"


def parse_native_discount_policy_workbook(content: bytes, segment_code: str | None = None):
    workbook = load_workbook(BytesIO(content), read_only=False, data_only=True)
    if _is_internal_template(workbook, _DISCOUNT_POLICY_KEY):
        raise NativeWorkbookNotRecognized
    if len(workbook.sheetnames) != 1:
        raise NativeWorkbookNotRecognized
    sheet = workbook[workbook.sheetnames[0]]
    if "BOOKING PROTECTION" not in _clean(sheet.cell(2, 3).value).upper():
        raise NativeWorkbookNotRecognized

    source_wef = _extract_wef(workbook)
    rows: list[tuple[int, dict[str, Any]]] = []
    row_no_out = 2

    for source_row in range(3, min(sheet.max_row, 19) + 1):
        scope_key = _clean(sheet.cell(source_row, 2).value)
        if not scope_key:
            continue
        inferred_segment = _policy_segment(scope_key)
        booking = _clean(sheet.cell(source_row, 3).value)
        buffer = sheet.cell(source_row, 4).value
        insurance = sheet.cell(source_row, 5).value
        if booking:
            days = _policy_number(booking, r"(\d+(?:\.\d+)?)")
            if days is not None:
                rows.append(
                    (
                        row_no_out,
                        {
                            "scope_type": "MODEL",
                            "segment_code": inferred_segment,
                            "scope_key": scope_key,
                            "parameter_key": "BOOKING_PROTECTION_DAYS",
                            "value_type": "NUMBER",
                            "value_number": str(days),
                            "value_text": None,
                            "unit": "DAY",
                            "effective_to": None,
                            "notes": None,
                            "source_effective_from": source_wef,
                            "source_row": source_row,
                        },
                    )
                )
                row_no_out += 1
        buffer_text = _clean(buffer)
        if buffer_text:
            amount = Decimal(0) if buffer_text.upper() == "NIL" else _number(buffer)
            if amount is not None:
                rows.append(
                    (
                        row_no_out,
                        {
                            "scope_type": "MODEL",
                            "segment_code": inferred_segment,
                            "scope_key": scope_key,
                            "parameter_key": "AGREED_BUFFER",
                            "value_type": "NUMBER",
                            "value_number": str(amount),
                            "value_text": None,
                            "unit": "INR",
                            "effective_to": None,
                            "notes": None,
                            "source_effective_from": source_wef,
                            "source_row": source_row,
                        },
                    )
                )
                row_no_out += 1
        if insurance not in (None, ""):
            pct = _number(insurance)
            if pct is not None:
                if pct <= 1:
                    pct *= 100
                rows.append(
                    (
                        row_no_out,
                        {
                            "scope_type": "MODEL",
                            "segment_code": inferred_segment,
                            "scope_key": scope_key,
                            "parameter_key": "INSURANCE_OD_PERCENT",
                            "value_type": "NUMBER",
                            "value_number": str(pct),
                            "value_text": None,
                            "unit": "PERCENT",
                            "effective_to": None,
                            "notes": None,
                            "source_effective_from": source_wef,
                            "source_row": source_row,
                        },
                    )
                )
                row_no_out += 1

    for source_row in range(21, sheet.max_row + 1):
        parameter = _clean(sheet.cell(source_row, 2).value)
        notes = _clean(sheet.cell(source_row, 3).value)
        if not parameter or not notes:
            continue
        code = _code(parameter, 80)
        rows.append(
            (
                row_no_out,
                {
                    "scope_type": "PROJECT",
                    "segment_code": None,
                    "scope_key": None,
                    "parameter_key": f"{code}_POLICY",
                    "value_type": "TEXT",
                    "value_number": None,
                    "value_text": notes,
                    "unit": None,
                    "effective_to": None,
                    "notes": notes,
                    "source_effective_from": source_wef,
                    "source_row": source_row,
                },
            )
        )
        row_no_out += 1

        extracted: list[tuple[str, Decimal, str]] = []
        if code == "AGED_STOCK":
            value = _policy_number(notes, r"above\s+(\d+)\s+days")
            if value is not None:
                extracted.append(("AGED_STOCK_MAX_DAYS", value, "DAY"))
        elif code == "BULK_DEAL":
            value = _policy_number(notes, r"(\d+)\s+or\s+more\s+vehicles")
            if value is not None:
                extracted.append(("BULK_DEAL_MIN_QUANTITY", value, "VEHICLE"))
        elif code == "MANAGEMENT_REFERRAL_MR":
            value = _policy_number(notes, r"(\d+(?:\.\d+)?)%")
            if value is not None:
                extracted.append(("MR_MAX_PERCENT_PREVIOUS_MONTH_RETAIL", value, "PERCENT"))
        elif code == "PENALTY_AMOUNT":
            value = _policy_number(notes.replace("₹", ""), r"([\d,]+)")
            if value is not None:
                extracted.append(("POLICY_DEVIATION_PENALTY", value, "INR"))
        elif code == "TRADE_IN":
            hold = _policy_number(notes, r"within\s+(\d+)\s+days")
            pv = _policy_number(notes.replace("₹", ""), r"([\d,]+)\s+for\s+Passenger")
            cv = _policy_number(notes.replace("₹", ""), r"([\d,]+)\s+for\s+Commercial")
            if hold is not None:
                extracted.append(("TRADE_IN_MAX_HOLDING_DAYS", hold, "DAY"))
            if pv is not None:
                extracted.append(("TRADE_IN_MIN_PROFIT_PV", pv, "INR"))
            if cv is not None:
                extracted.append(("TRADE_IN_MIN_PROFIT_CV", cv, "INR"))
        for key, value, unit in extracted:
            rows.append(
                (
                    row_no_out,
                    {
                        "scope_type": "PROJECT",
                        "segment_code": None,
                        "scope_key": None,
                        "parameter_key": key,
                        "value_type": "NUMBER",
                        "value_number": str(value),
                        "value_text": None,
                        "unit": unit,
                        "effective_to": None,
                        "notes": notes,
                        "source_effective_from": source_wef,
                        "source_row": source_row,
                    },
                )
            )
            row_no_out += 1
    if not rows:
        raise NativeWorkbookNotRecognized
    return rows


def install_native_workbook_parser(mahindra_masters_module) -> None:
    original = mahindra_masters_module._workbook_rows

    def _workbook_rows(content: bytes, *, master_key: str, segment_code: str | None = None):
        try:
            if master_key == _SEGMENT_MASTER_KEY:
                return parse_native_segment_workbook(content, segment_code)
            if master_key == _DISCOUNT_POLICY_KEY:
                return parse_native_discount_policy_workbook(content, segment_code)
        except NativeWorkbookNotRecognized:
            pass
        return original(content, master_key=master_key, segment_code=segment_code)

    mahindra_masters_module._workbook_rows = _workbook_rows
