from __future__ import annotations

import hashlib
import json
import re
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import Annotated, Any
from uuid import UUID

from fastapi import APIRouter, Depends, File, Form, Header, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from pydantic import BaseModel
from sqlalchemy import Connection, text

from audit_core.db import set_tenant_context
from audit_core.dependencies import (
    HumanAdminRequest,
    get_connection,
    require_super_admin_request,
)
from audit_core.errors import ConflictError, NotFoundError, ValidationError
from audit_core.idempotency import stable_request_hash
from audit_core.price_lists import (
    create_price_list,
    create_price_list_version,
    publish_price_list_version,
)
from audit_core.product_masters import (
    add_project_product_master_item,
    create_project_product_master_version,
    publish_project_product_master_version,
)

router = APIRouter(prefix="/v1/tenants/{tenant_id}/mahindra-masters", tags=["mahindra-masters"])

_SEGMENT_MASTER_KEY = "MAHINDRA_SEGMENT_MASTER"
_DISCOUNT_POLICY_KEY = "DISCOUNT_POLICY"
_TEMPLATE_VERSION = "1.0"

_SEGMENT_COLUMNS = (
    "sku_code",
    "model_code",
    "model_name",
    "model_year",
    "trim_code",
    "trim_name",
    "configuration_code",
    "fuel_powertrain",
    "transmission",
    "drive_type",
    "seating_capacity",
    "body_type",
    "colour_code",
    "colour_name",
    "currency_code",
    "component_key",
    "component_label",
    "standard_amount",
)
_SEGMENT_REQUIRED = frozenset(
    {
        "sku_code",
        "model_code",
        "model_name",
        "trim_code",
        "trim_name",
        "configuration_code",
        "component_key",
        "standard_amount",
    }
)

_POLICY_COLUMNS = (
    "scope_type",
    "segment_code",
    "scope_key",
    "parameter_key",
    "value_type",
    "value_number",
    "value_text",
    "unit",
    "effective_to",
    "notes",
)
_POLICY_REQUIRED = frozenset({"scope_type", "parameter_key", "value_type"})
_POLICY_SCOPE_TYPES = frozenset({"PROJECT", "SEGMENT", "MODEL", "TRIM", "CONFIGURATION"})
_POLICY_VALUE_TYPES = frozenset({"NUMBER", "TEXT", "BOOLEAN"})


class MahindraImportResponse(BaseModel):
    importId: UUID
    masterKey: str
    segmentId: UUID | None
    segmentCode: str | None
    effectiveFrom: date
    fileName: str
    status: str
    rowsParsed: int
    validRows: int
    errorRows: int
    productMasterVersionId: UUID | None = None
    priceListVersionId: UUID | None = None
    discountPolicyVersionId: UUID | None = None
    lifecycleStatus: str | None = None


class MahindraMasterOption(BaseModel):
    segmentId: UUID | None
    segmentCode: str | None
    segmentName: str | None
    uploadKey: str
    displayName: str


class MahindraMasterOptionsResponse(BaseModel):
    oemCode: str
    segmentUploads: list[MahindraMasterOption]
    discountUpload: MahindraMasterOption


def _normalise_cell(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, Decimal):
        return str(value)
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return value


def _text_value(row: dict[str, Any], key: str) -> str | None:
    value = row.get(key)
    if value is None or not str(value).strip():
        return None
    return str(value).strip()


def _decimal(value: Any) -> Decimal | None:
    if value is None or str(value).strip() == "":
        return None
    try:
        return Decimal(str(value).replace(",", "").strip())
    except InvalidOperation:
        return None


def _integer(value: Any) -> int | None:
    if value is None or str(value).strip() == "":
        return None
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def _date(value: Any) -> date | None:
    if value is None or str(value).strip() == "":
        return None
    if isinstance(value, date):
        return value
    try:
        return date.fromisoformat(str(value).strip())
    except ValueError:
        return None


def _component_key(value: str) -> str:
    normalized = re.sub(r"[^A-Z0-9]+", "_", value.strip().upper()).strip("_")
    if not normalized:
        raise ValidationError(detail="Price component key cannot be empty.")
    return normalized[:100]


def _project_context(connection: Connection, tenant_id: str):
    row = connection.execute(
        text(
            """
            SELECT p.oem_id, o.oem_code, o.oem_name
            FROM auditcore.projects p
            JOIN auditcore.oems o ON o.oem_id=p.oem_id
            WHERE p.tenant_id=:tenant_id
            """
        ),
        {"tenant_id": tenant_id},
    ).mappings().one_or_none()
    if row is None:
        raise NotFoundError(
            error_code="VAC-NF-001",
            title="Project not found",
            detail="Project not found for the requested tenant.",
        )
    if str(row["oem_code"]).upper() != "MAHINDRA":
        raise ValidationError(detail="This upload path is available only for Mahindra OEM Projects.")
    return row


def _selected_segments(connection: Connection, tenant_id: str):
    return connection.execute(
        text(
            """
            SELECT s.segment_id, s.segment_code, s.segment_name
            FROM auditcore.project_segments ps
            JOIN auditcore.oem_segments s ON s.segment_id=ps.segment_id
            WHERE ps.tenant_id=:tenant_id AND s.is_active=true
            ORDER BY s.segment_name, s.segment_code
            """
        ),
        {"tenant_id": tenant_id},
    ).mappings().all()


def _require_selected_segment(connection: Connection, tenant_id: str, segment_id: UUID):
    row = connection.execute(
        text(
            """
            SELECT s.segment_id, s.segment_code, s.segment_name
            FROM auditcore.project_segments ps
            JOIN auditcore.oem_segments s ON s.segment_id=ps.segment_id
            JOIN auditcore.projects p ON p.tenant_id=ps.tenant_id AND p.oem_id=s.oem_id
            JOIN auditcore.oems o ON o.oem_id=p.oem_id
            WHERE ps.tenant_id=:tenant_id
              AND ps.segment_id=:segment_id
              AND s.is_active=true
              AND o.oem_code='MAHINDRA'
            """
        ),
        {"tenant_id": tenant_id, "segment_id": segment_id},
    ).mappings().one_or_none()
    if row is None:
        raise ValidationError(detail="Segment is not selected for this Mahindra Project.")
    return row


def _build_template(*, master_key: str, segment_code: str | None = None) -> bytes:
    columns = _SEGMENT_COLUMNS if master_key == _SEGMENT_MASTER_KEY else _POLICY_COLUMNS
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "VEHICLE_PRICE" if master_key == _SEGMENT_MASTER_KEY else "DISCOUNT_POLICY"
    for index, column in enumerate(columns, start=1):
        sheet.cell(row=1, column=index, value=column)
    meta = workbook.create_sheet("_meta")
    meta["A1"] = "master_key"
    meta["B1"] = master_key
    meta["A2"] = "template_version"
    meta["B2"] = _TEMPLATE_VERSION
    if segment_code:
        meta["A3"] = "segment_code"
        meta["B3"] = segment_code
    meta.sheet_state = "hidden"
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _workbook_rows(content: bytes, *, master_key: str, segment_code: str | None = None):
    try:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise ValidationError(detail="The uploaded .xlsx workbook is unreadable.") from exc
    if "_meta" in workbook.sheetnames:
        meta = workbook["_meta"]
        if meta["B1"].value != master_key:
            raise ValidationError(detail="Workbook master metadata does not match this upload.")
        if str(meta["B2"].value) != _TEMPLATE_VERSION:
            raise ValidationError(detail="Workbook template version is not supported.")
        if segment_code and meta["B3"].value and str(meta["B3"].value) != segment_code:
            raise ValidationError(detail="Workbook Segment does not match the selected Segment.")

    data_sheets = [name for name in workbook.sheetnames if name != "_meta"]
    if len(data_sheets) != 1:
        raise ValidationError(detail="Master workbook must contain exactly one data sheet.")
    sheet = workbook[data_sheets[0]]
    iterator = sheet.iter_rows(values_only=True)
    try:
        raw_headers = next(iterator)
    except StopIteration as exc:
        raise ValidationError(detail="Master workbook is empty.") from exc
    headers = [str(value).strip() if value is not None else "" for value in raw_headers]
    if not headers or any(not header for header in headers):
        raise ValidationError(detail="Master workbook header contains an empty column name.")
    if len(headers) != len(set(headers)):
        raise ValidationError(detail="Master workbook header contains duplicate columns.")

    allowed = set(_SEGMENT_COLUMNS if master_key == _SEGMENT_MASTER_KEY else _POLICY_COLUMNS)
    required = _SEGMENT_REQUIRED if master_key == _SEGMENT_MASTER_KEY else _POLICY_REQUIRED
    unknown = sorted(set(headers) - allowed)
    missing = sorted(required - set(headers))
    if unknown:
        raise ValidationError(detail=f"Workbook contains unsupported column(s): {', '.join(unknown)}.")
    if missing:
        raise ValidationError(detail=f"Workbook is missing required column(s): {', '.join(missing)}.")

    rows: list[tuple[int, dict[str, Any]]] = []
    for row_number, values in enumerate(iterator, start=2):
        if all(value is None or str(value).strip() == "" for value in values):
            continue
        parsed = {
            header: _normalise_cell(value)
            for header, value in zip(headers, values, strict=False)
        }
        rows.append((row_number, parsed))
    if not rows:
        raise ValidationError(detail="Master workbook contains no data rows.")
    return rows


def _validate_segment_rows(rows: list[tuple[int, dict[str, Any]]]):
    validated: list[tuple[int, dict[str, Any], list[str]]] = []
    sku_identity: dict[str, tuple[object, ...]] = {}
    sku_components: set[tuple[str, str]] = set()
    for row_number, row in rows:
        messages: list[str] = []
        for key in _SEGMENT_REQUIRED:
            if _text_value(row, key) is None:
                messages.append(f"{key} is required.")
        model_year = _integer(row.get("model_year"))
        if row.get("model_year") not in (None, "") and model_year is None:
            messages.append("model_year must be a whole number.")
        seats = _integer(row.get("seating_capacity"))
        if row.get("seating_capacity") not in (None, "") and (seats is None or seats <= 0):
            messages.append("seating_capacity must be a positive whole number.")
        amount = _decimal(row.get("standard_amount"))
        if amount is None:
            messages.append("standard_amount must be numeric.")
        currency = (_text_value(row, "currency_code") or "INR").upper()
        if len(currency) != 3:
            messages.append("currency_code must be a three-letter code.")

        colour_code = _text_value(row, "colour_code")
        colour_name = _text_value(row, "colour_name")
        if colour_code and not colour_name:
            messages.append("colour_name is required when colour_code is supplied.")
        if colour_name and not colour_code:
            messages.append("colour_code is required when colour_name is supplied.")

        sku_code = _text_value(row, "sku_code")
        if sku_code:
            identity = (
                _text_value(row, "model_code"),
                _text_value(row, "trim_code"),
                _text_value(row, "configuration_code"),
                _text_value(row, "fuel_powertrain"),
                _text_value(row, "transmission"),
                _text_value(row, "drive_type"),
                seats,
                colour_code,
            )
            previous = sku_identity.get(sku_code)
            if previous is not None and previous != identity:
                messages.append("sku_code is repeated with a different vehicle configuration.")
            else:
                sku_identity[sku_code] = identity

            component = _text_value(row, "component_key")
            if component:
                normalized_component = _component_key(component)
                pair = (sku_code, normalized_component)
                if pair in sku_components:
                    messages.append("The same SKU/component appears more than once in this upload.")
                else:
                    sku_components.add(pair)
                row["component_key"] = normalized_component

        row["currency_code"] = currency
        if amount is not None:
            row["standard_amount"] = str(amount)
        if model_year is not None:
            row["model_year"] = model_year
        if seats is not None:
            row["seating_capacity"] = seats
        validated.append((row_number, row, messages))
    return validated


def _validate_policy_rows(connection: Connection, tenant_id: str, rows):
    selected_by_code = {
        str(row["segment_code"]): row for row in _selected_segments(connection, tenant_id)
    }
    validated: list[tuple[int, dict[str, Any], list[str]]] = []
    seen: set[tuple[str, str | None, str | None, str]] = set()
    for row_number, row in rows:
        messages: list[str] = []
        scope_type = (_text_value(row, "scope_type") or "").upper()
        value_type = (_text_value(row, "value_type") or "").upper()
        parameter_key = _text_value(row, "parameter_key")
        segment_code = _text_value(row, "segment_code")
        scope_key = _text_value(row, "scope_key")
        if scope_type not in _POLICY_SCOPE_TYPES:
            messages.append("scope_type must be PROJECT, SEGMENT, MODEL, TRIM or CONFIGURATION.")
        if value_type not in _POLICY_VALUE_TYPES:
            messages.append("value_type must be NUMBER, TEXT or BOOLEAN.")
        if not parameter_key:
            messages.append("parameter_key is required.")
        if scope_type != "PROJECT" and not segment_code:
            messages.append("segment_code is required for non-PROJECT policy scope.")
        if segment_code and segment_code not in selected_by_code:
            messages.append("segment_code must be one of the Segments selected for this Project.")
        if scope_type in {"MODEL", "TRIM", "CONFIGURATION"} and not scope_key:
            messages.append("scope_key is required for MODEL, TRIM and CONFIGURATION scope.")

        value_number = _decimal(row.get("value_number"))
        value_text = _text_value(row, "value_text")
        if value_type == "NUMBER":
            if value_number is None:
                messages.append("value_number is required and must be numeric for NUMBER parameters.")
            if value_text:
                messages.append("value_text must be empty for NUMBER parameters.")
        elif value_type == "TEXT":
            if not value_text:
                messages.append("value_text is required for TEXT parameters.")
            if value_number is not None:
                messages.append("value_number must be empty for TEXT parameters.")
        elif value_type == "BOOLEAN":
            if not value_text or value_text.lower() not in {"true", "false", "yes", "no", "1", "0"}:
                messages.append("value_text must be true/false for BOOLEAN parameters.")
            else:
                row["value_text"] = "true" if value_text.lower() in {"true", "yes", "1"} else "false"
            if value_number is not None:
                messages.append("value_number must be empty for BOOLEAN parameters.")

        effective_to = _date(row.get("effective_to"))
        if row.get("effective_to") not in (None, "") and effective_to is None:
            messages.append("effective_to must be YYYY-MM-DD.")
        if effective_to is not None:
            row["effective_to"] = effective_to.isoformat()
        row["scope_type"] = scope_type
        row["value_type"] = value_type
        if value_number is not None:
            row["value_number"] = str(value_number)

        if parameter_key:
            dedupe = (scope_type, segment_code, scope_key, parameter_key.upper())
            if dedupe in seen:
                messages.append("The same policy parameter/scope appears more than once in this upload.")
            else:
                seen.add(dedupe)
            row["parameter_key"] = parameter_key.upper()
        validated.append((row_number, row, messages))
    return validated


def _stage_import(
    connection: Connection,
    *,
    tenant_id: str,
    master_key: str,
    segment_id: UUID | None,
    effective_from: date,
    file_name: str,
    file_content: bytes,
    idempotency_key: str,
    actor_id: str,
    validated_rows,
):
    file_hash = hashlib.sha256(file_content).hexdigest()
    semantic_payload = {
        "tenantId": tenant_id,
        "masterKey": master_key,
        "segmentId": str(segment_id) if segment_id else None,
        "effectiveFrom": effective_from.isoformat(),
        "fileHash": file_hash,
    }
    request_hash = stable_request_hash(semantic_payload)
    existing = connection.execute(
        text(
            """
            SELECT import_id, semantic_request_hash
            FROM auditcore.project_master_imports
            WHERE tenant_id=:tenant_id AND owner_module='AUDIT_CORE'
              AND master_key=:master_key AND idempotency_key=:idempotency_key
            """
        ),
        {
            "tenant_id": tenant_id,
            "master_key": master_key,
            "idempotency_key": idempotency_key,
        },
    ).mappings().one_or_none()
    if existing is not None:
        if existing["semantic_request_hash"] != request_hash:
            raise ConflictError(
                error_code="VAC-CONFLICT-003",
                title="Idempotency conflict",
                detail="The Idempotency-Key was already used with a different master upload.",
            )
        return existing["import_id"]

    valid_rows = sum(1 for _, _, messages in validated_rows if not messages)
    error_rows = len(validated_rows) - valid_rows
    status = "PREVIEW_READY" if error_rows == 0 else "VALIDATION_FAILED"
    import_id = connection.execute(
        text(
            """
            INSERT INTO auditcore.project_master_imports (
                tenant_id, owner_module, master_key, segment_id, effective_from,
                template_version, original_file_name, file_hash, idempotency_key,
                semantic_request_hash, status, rows_parsed, valid_rows, warning_rows,
                error_rows, created_by_user_id
            ) VALUES (
                :tenant_id, 'AUDIT_CORE', :master_key, :segment_id, :effective_from,
                :template_version, :file_name, :file_hash, :idempotency_key,
                :request_hash, :status, :rows_parsed, :valid_rows, 0,
                :error_rows, :actor_id
            ) RETURNING import_id
            """
        ),
        {
            "tenant_id": tenant_id,
            "master_key": master_key,
            "segment_id": segment_id,
            "effective_from": effective_from,
            "template_version": _TEMPLATE_VERSION,
            "file_name": file_name,
            "file_hash": file_hash,
            "idempotency_key": idempotency_key,
            "request_hash": request_hash,
            "status": status,
            "rows_parsed": len(validated_rows),
            "valid_rows": valid_rows,
            "error_rows": error_rows,
            "actor_id": actor_id,
        },
    ).scalar_one()
    for row_number, row, messages in validated_rows:
        connection.execute(
            text(
                """
                INSERT INTO auditcore.project_master_import_rows (
                    tenant_id, import_id, row_number, parsed_data,
                    validation_status, validation_messages
                ) VALUES (
                    :tenant_id, :import_id, :row_number, CAST(:parsed_data AS jsonb),
                    :validation_status, CAST(:messages AS jsonb)
                )
                """
            ),
            {
                "tenant_id": tenant_id,
                "import_id": import_id,
                "row_number": row_number,
                "parsed_data": json.dumps(row, default=str),
                "validation_status": "ERROR" if messages else "VALID",
                "messages": json.dumps(messages),
            },
        )
    return import_id


def _import_record(connection: Connection, tenant_id: str, import_id: UUID):
    row = connection.execute(
        text(
            """
            SELECT i.import_id, i.master_key, i.segment_id, i.effective_from,
                   i.original_file_name, i.status, i.rows_parsed, i.valid_rows,
                   i.error_rows, i.confirmation_receipt,
                   s.segment_code
            FROM auditcore.project_master_imports i
            LEFT JOIN auditcore.oem_segments s ON s.segment_id=i.segment_id
            WHERE i.tenant_id=:tenant_id AND i.import_id=:import_id
            """
        ),
        {"tenant_id": tenant_id, "import_id": import_id},
    ).mappings().one_or_none()
    if row is None:
        raise NotFoundError(
            error_code="VAC-NF-008",
            title="Project Master import not found",
            detail="The requested Mahindra master import does not exist for this Project.",
        )
    return row


def _response(row, *, lifecycle_status: str | None = None) -> MahindraImportResponse:
    receipt = dict(row["confirmation_receipt"] or {})
    return MahindraImportResponse(
        importId=row["import_id"],
        masterKey=row["master_key"],
        segmentId=row["segment_id"],
        segmentCode=row["segment_code"],
        effectiveFrom=row["effective_from"],
        fileName=row["original_file_name"],
        status=row["status"],
        rowsParsed=int(row["rows_parsed"]),
        validRows=int(row["valid_rows"]),
        errorRows=int(row["error_rows"]),
        productMasterVersionId=receipt.get("productMasterVersionId"),
        priceListVersionId=receipt.get("priceListVersionId"),
        discountPolicyVersionId=receipt.get("discountPolicyVersionId"),
        lifecycleStatus=lifecycle_status,
    )


def _staged_rows(connection: Connection, tenant_id: str, import_id: UUID):
    return [
        dict(row["parsed_data"])
        for row in connection.execute(
            text(
                """
                SELECT parsed_data
                FROM auditcore.project_master_import_rows
                WHERE tenant_id=:tenant_id AND import_id=:import_id
                ORDER BY row_number
                """
            ),
            {"tenant_id": tenant_id, "import_id": import_id},
        ).mappings().all()
    ]


def _ensure_model(connection: Connection, *, oem_id: UUID, segment_id: UUID, row):
    code = str(row["model_code"]).strip()
    name = str(row["model_name"]).strip()
    current = connection.execute(
        text(
            """
            SELECT model_id, model_name, segment_id
            FROM auditcore.product_models
            WHERE oem_id=:oem_id AND model_code=:code
            """
        ),
        {"oem_id": oem_id, "code": code},
    ).mappings().one_or_none()
    if current is not None:
        if current["model_name"] != name:
            raise ConflictError(
                error_code="VAC-CONFLICT-004",
                title="Canonical Product conflict",
                detail=f"Model {code} already exists with a different name.",
            )
        if current["segment_id"] not in (None, segment_id):
            raise ConflictError(
                error_code="VAC-CONFLICT-004",
                title="Canonical Product conflict",
                detail=f"Model {code} already belongs to a different OEM Segment.",
            )
        if current["segment_id"] is None:
            connection.execute(
                text("UPDATE auditcore.product_models SET segment_id=:segment_id WHERE model_id=:id"),
                {"segment_id": segment_id, "id": current["model_id"]},
            )
        return current["model_id"]
    return connection.execute(
        text(
            """
            INSERT INTO auditcore.product_models (
                oem_id, segment_id, model_code, model_name, model_year
            ) VALUES (:oem_id, :segment_id, :code, :name, :model_year)
            RETURNING model_id
            """
        ),
        {
            "oem_id": oem_id,
            "segment_id": segment_id,
            "code": code,
            "name": name,
            "model_year": row.get("model_year"),
        },
    ).scalar_one()


def _ensure_trim(connection: Connection, *, model_id: UUID, row):
    code = str(row["trim_code"]).strip()
    name = str(row["trim_name"]).strip()
    current = connection.execute(
        text(
            "SELECT variant_id, variant_name FROM auditcore.product_variants "
            "WHERE model_id=:model_id AND variant_code=:code"
        ),
        {"model_id": model_id, "code": code},
    ).mappings().one_or_none()
    if current is not None:
        if current["variant_name"] != name:
            raise ConflictError(
                error_code="VAC-CONFLICT-004",
                title="Canonical Product conflict",
                detail=f"Trim {code} already exists with a different name.",
            )
        return current["variant_id"]
    return connection.execute(
        text(
            """
            INSERT INTO auditcore.product_variants (
                model_id, variant_code, variant_name, body_type
            ) VALUES (:model_id, :code, :name, :body_type)
            RETURNING variant_id
            """
        ),
        {
            "model_id": model_id,
            "code": code,
            "name": name,
            "body_type": _text_value(row, "body_type"),
        },
    ).scalar_one()


def _ensure_configuration(connection: Connection, *, segment_id: UUID, model_id: UUID, trim_id: UUID, row):
    code = str(row["configuration_code"]).strip()
    attrs = {
        "source": "MAHINDRA_SEGMENT_MASTER",
    }
    expected = {
        "fuel_powertrain": _text_value(row, "fuel_powertrain"),
        "transmission": _text_value(row, "transmission"),
        "drive_type": _text_value(row, "drive_type"),
        "seating_capacity": _integer(row.get("seating_capacity")),
        "body_type": _text_value(row, "body_type"),
    }
    current = connection.execute(
        text(
            """
            SELECT configuration_id, segment_id, model_id, variant_id,
                   fuel_powertrain, transmission, drive_type, seating_capacity, body_type
            FROM auditcore.product_configurations
            WHERE variant_id=:trim_id AND configuration_code=:code
            """
        ),
        {"trim_id": trim_id, "code": code},
    ).mappings().one_or_none()
    if current is not None:
        if (
            current["segment_id"] != segment_id
            or current["model_id"] != model_id
            or any(current[key] != value for key, value in expected.items())
        ):
            raise ConflictError(
                error_code="VAC-CONFLICT-004",
                title="Canonical Product conflict",
                detail=f"Configuration {code} already identifies different vehicle attributes.",
            )
        return current["configuration_id"]
    return connection.execute(
        text(
            """
            INSERT INTO auditcore.product_configurations (
                segment_id, model_id, variant_id, configuration_code,
                fuel_powertrain, transmission, drive_type, seating_capacity,
                body_type, attributes
            ) VALUES (
                :segment_id, :model_id, :trim_id, :code,
                :fuel, :transmission, :drive_type, :seating_capacity,
                :body_type, CAST(:attributes AS jsonb)
            ) RETURNING configuration_id
            """
        ),
        {
            "segment_id": segment_id,
            "model_id": model_id,
            "trim_id": trim_id,
            "code": code,
            "fuel": expected["fuel_powertrain"],
            "transmission": expected["transmission"],
            "drive_type": expected["drive_type"],
            "seating_capacity": expected["seating_capacity"],
            "body_type": expected["body_type"],
            "attributes": json.dumps(attrs),
        },
    ).scalar_one()


def _ensure_colour(connection: Connection, *, oem_id: UUID, row):
    code = _text_value(row, "colour_code")
    name = _text_value(row, "colour_name")
    if not code:
        return None
    current = connection.execute(
        text(
            "SELECT colour_id, colour_name FROM auditcore.colours "
            "WHERE oem_id=:oem_id AND colour_code=:code"
        ),
        {"oem_id": oem_id, "code": code},
    ).mappings().one_or_none()
    if current is not None:
        if current["colour_name"] != name:
            raise ConflictError(
                error_code="VAC-CONFLICT-004",
                title="Canonical Product conflict",
                detail=f"Colour {code} already exists with a different name.",
            )
        return current["colour_id"]
    return connection.execute(
        text(
            "INSERT INTO auditcore.colours (oem_id, colour_code, colour_name) "
            "VALUES (:oem_id, :code, :name) RETURNING colour_id"
        ),
        {"oem_id": oem_id, "code": code, "name": name},
    ).scalar_one()


def _ensure_sku(
    connection: Connection,
    *,
    oem_id: UUID,
    model_id: UUID,
    trim_id: UUID,
    configuration_id: UUID,
    colour_id: UUID | None,
    row,
):
    sku_code = str(row["sku_code"]).strip()
    current = connection.execute(
        text(
            """
            SELECT product_sku_id, model_id, variant_id, configuration_id, colour_id
            FROM auditcore.product_skus
            WHERE oem_id=:oem_id AND sku_code=:sku_code
            """
        ),
        {"oem_id": oem_id, "sku_code": sku_code},
    ).mappings().one_or_none()
    if current is not None:
        if (
            current["model_id"] != model_id
            or current["variant_id"] != trim_id
            or current["configuration_id"] not in (None, configuration_id)
            or current["colour_id"] != colour_id
        ):
            raise ConflictError(
                error_code="VAC-CONFLICT-004",
                title="Canonical Product conflict",
                detail=f"SKU {sku_code} already identifies a different vehicle configuration.",
            )
        if current["configuration_id"] is None:
            connection.execute(
                text(
                    "UPDATE auditcore.product_skus SET configuration_id=:configuration_id "
                    "WHERE product_sku_id=:sku_id"
                ),
                {"configuration_id": configuration_id, "sku_id": current["product_sku_id"]},
            )
        return current["product_sku_id"]
    return connection.execute(
        text(
            """
            INSERT INTO auditcore.product_skus (
                oem_id, model_id, variant_id, configuration_id, colour_id, sku_code, attributes
            ) VALUES (
                :oem_id, :model_id, :trim_id, :configuration_id, :colour_id, :sku_code,
                jsonb_build_object('source', 'MAHINDRA_SEGMENT_MASTER')
            ) RETURNING product_sku_id
            """
        ),
        {
            "oem_id": oem_id,
            "model_id": model_id,
            "trim_id": trim_id,
            "configuration_id": configuration_id,
            "colour_id": colour_id,
            "sku_code": sku_code,
        },
    ).scalar_one()


def _ensure_segment_price_list(connection: Connection, *, tenant_id: str, segment_code: str, segment_name: str, actor_id: str):
    code = f"MAHINDRA_{segment_code}"[:120]
    existing = connection.execute(
        text(
            "SELECT price_list_id FROM auditcore.price_lists "
            "WHERE tenant_id=:tenant_id AND price_list_code=:code"
        ),
        {"tenant_id": tenant_id, "code": code},
    ).scalar_one_or_none()
    if existing is not None:
        return existing
    return create_price_list(
        connection,
        tenant_id=tenant_id,
        code=code,
        name=f"Mahindra {segment_name} Price Master",
        actor_id=actor_id,
    )


def _confirm_segment_master(connection: Connection, *, tenant_id: str, import_row, actor_id: str):
    segment = _require_selected_segment(connection, tenant_id, import_row["segment_id"])
    project = _project_context(connection, tenant_id)
    rows = _staged_rows(connection, tenant_id, import_row["import_id"])

    product_version_id = create_project_product_master_version(
        connection,
        tenant_id=tenant_id,
        segment_id=segment["segment_id"],
        effective_from=import_row["effective_from"],
        actor_id=actor_id,
        source_import_id=import_row["import_id"],
    )
    price_list_id = _ensure_segment_price_list(
        connection,
        tenant_id=tenant_id,
        segment_code=segment["segment_code"],
        segment_name=segment["segment_name"],
        actor_id=actor_id,
    )
    version_no = connection.execute(
        text(
            "SELECT COALESCE(MAX(version_no),0)+1 FROM auditcore.price_list_versions "
            "WHERE tenant_id=:tenant_id AND price_list_id=:price_list_id"
        ),
        {"tenant_id": tenant_id, "price_list_id": price_list_id},
    ).scalar_one()
    price_version_id = create_price_list_version(
        connection,
        tenant_id=tenant_id,
        price_list_id=price_list_id,
        version_no=int(version_no),
        effective_from=import_row["effective_from"],
        actor_id=actor_id,
    )

    sku_ids: dict[str, UUID] = {}
    added_master_skus: set[UUID] = set()
    currencies = {str(row.get("currency_code") or "INR").upper() for row in rows}
    if len(currencies) > 1:
        raise ValidationError(detail="One Segment Price Master version must use one currency.")
    currency_code = next(iter(currencies), "INR")
    connection.execute(
        text(
            "UPDATE auditcore.price_list_versions SET currency_code=:currency "
            "WHERE tenant_id=:tenant_id AND price_list_version_id=:version_id"
        ),
        {"currency": currency_code, "tenant_id": tenant_id, "version_id": price_version_id},
    )

    for row in rows:
        sku_code = str(row["sku_code"]).strip()
        sku_id = sku_ids.get(sku_code)
        if sku_id is None:
            model_id = _ensure_model(
                connection,
                oem_id=project["oem_id"],
                segment_id=segment["segment_id"],
                row=row,
            )
            trim_id = _ensure_trim(connection, model_id=model_id, row=row)
            configuration_id = _ensure_configuration(
                connection,
                segment_id=segment["segment_id"],
                model_id=model_id,
                trim_id=trim_id,
                row=row,
            )
            colour_id = _ensure_colour(connection, oem_id=project["oem_id"], row=row)
            sku_id = _ensure_sku(
                connection,
                oem_id=project["oem_id"],
                model_id=model_id,
                trim_id=trim_id,
                configuration_id=configuration_id,
                colour_id=colour_id,
                row=row,
            )
            sku_ids[sku_code] = sku_id
        if sku_id not in added_master_skus:
            add_project_product_master_item(
                connection,
                tenant_id=tenant_id,
                version_id=product_version_id,
                product_sku_id=sku_id,
            )
            added_master_skus.add(sku_id)

        component_key = _component_key(str(row["component_key"]))
        metadata = {
            "componentLabel": _text_value(row, "component_label") or component_key,
            "source": "MAHINDRA_SEGMENT_MASTER",
            "segmentCode": segment["segment_code"],
        }
        connection.execute(
            text(
                """
                INSERT INTO auditcore.price_list_items (
                    tenant_id, price_list_version_id, product_sku_id,
                    component_key, standard_amount, metadata
                ) VALUES (
                    :tenant_id, :version_id, :sku_id,
                    :component_key, :amount, CAST(:metadata AS jsonb)
                )
                """
            ),
            {
                "tenant_id": tenant_id,
                "version_id": price_version_id,
                "sku_id": sku_id,
                "component_key": component_key,
                "amount": Decimal(str(row["standard_amount"])),
                "metadata": json.dumps(metadata),
            },
        )

    return product_version_id, price_version_id


def _confirm_discount_policy(connection: Connection, *, tenant_id: str, import_row, actor_id: str):
    rows = _staged_rows(connection, tenant_id, import_row["import_id"])
    selected_by_code = {
        str(row["segment_code"]): row for row in _selected_segments(connection, tenant_id)
    }
    version_no = int(
        connection.execute(
            text(
                "SELECT COALESCE(MAX(version_no),0)+1 FROM auditcore.discount_policy_versions "
                "WHERE tenant_id=:tenant_id"
            ),
            {"tenant_id": tenant_id},
        ).scalar_one()
    )
    effective_to_values = {
        _date(row.get("effective_to")) for row in rows if row.get("effective_to") not in (None, "")
    }
    if len(effective_to_values) > 1:
        raise ValidationError(detail="One Discount & Policy version must use one Effective To date.")
    effective_to = next(iter(effective_to_values), None)
    version_id = connection.execute(
        text(
            """
            INSERT INTO auditcore.discount_policy_versions (
                tenant_id, version_no, effective_from, effective_to,
                source_import_id, created_by_actor_id
            ) VALUES (
                :tenant_id, :version_no, :effective_from, :effective_to,
                :source_import_id, :actor_id
            ) RETURNING discount_policy_version_id
            """
        ),
        {
            "tenant_id": tenant_id,
            "version_no": version_no,
            "effective_from": import_row["effective_from"],
            "effective_to": effective_to,
            "source_import_id": import_row["import_id"],
            "actor_id": actor_id,
        },
    ).scalar_one()

    for index, row in enumerate(rows, start=2):
        segment_code = _text_value(row, "segment_code")
        segment_id = selected_by_code[segment_code]["segment_id"] if segment_code else None
        value_type = str(row["value_type"]).upper()
        value_number = Decimal(str(row["value_number"])) if value_type == "NUMBER" else None
        value_text = str(row["value_text"]) if value_type in {"TEXT", "BOOLEAN"} else None
        connection.execute(
            text(
                """
                INSERT INTO auditcore.discount_policy_parameters (
                    tenant_id, discount_policy_version_id, scope_type,
                    segment_id, scope_key, parameter_key, value_type,
                    value_number, value_text, unit, notes, source_import_row_no
                ) VALUES (
                    :tenant_id, :version_id, :scope_type,
                    :segment_id, :scope_key, :parameter_key, :value_type,
                    :value_number, :value_text, :unit, :notes, :row_no
                )
                """
            ),
            {
                "tenant_id": tenant_id,
                "version_id": version_id,
                "scope_type": row["scope_type"],
                "segment_id": segment_id,
                "scope_key": _text_value(row, "scope_key"),
                "parameter_key": row["parameter_key"],
                "value_type": value_type,
                "value_number": value_number,
                "value_text": value_text,
                "unit": _text_value(row, "unit"),
                "notes": _text_value(row, "notes"),
                "row_no": index,
            },
        )
    return version_id


@router.get("/options", response_model=MahindraMasterOptionsResponse)
def get_options(
    tenant_id: str,
    admin_request: Annotated[HumanAdminRequest, Depends(require_super_admin_request)],
    connection: Annotated[Connection, Depends(get_connection)],
) -> MahindraMasterOptionsResponse:
    del admin_request
    set_tenant_context(connection, tenant_id)
    _project_context(connection, tenant_id)
    segments = _selected_segments(connection, tenant_id)
    return MahindraMasterOptionsResponse(
        oemCode="MAHINDRA",
        segmentUploads=[
            MahindraMasterOption(
                segmentId=row["segment_id"],
                segmentCode=row["segment_code"],
                segmentName=row["segment_name"],
                uploadKey=_SEGMENT_MASTER_KEY,
                displayName=f"{row['segment_name']} Vehicle & Price Master",
            )
            for row in segments
        ],
        discountUpload=MahindraMasterOption(
            segmentId=None,
            segmentCode=None,
            segmentName=None,
            uploadKey=_DISCOUNT_POLICY_KEY,
            displayName="Discount & Policy Master",
        ),
    )


@router.get("/segments/{segment_id}/template")
def download_segment_template(
    tenant_id: str,
    segment_id: UUID,
    admin_request: Annotated[HumanAdminRequest, Depends(require_super_admin_request)],
    connection: Annotated[Connection, Depends(get_connection)],
):
    del admin_request
    set_tenant_context(connection, tenant_id)
    segment = _require_selected_segment(connection, tenant_id, segment_id)
    content = _build_template(master_key=_SEGMENT_MASTER_KEY, segment_code=segment["segment_code"])
    return StreamingResponse(
        BytesIO(content),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": (
                f'attachment; filename="mahindra-{str(segment["segment_code"]).lower()}-vehicle-price.xlsx"'
            )
        },
    )


@router.get("/discount-policy/template")
def download_discount_policy_template(
    tenant_id: str,
    admin_request: Annotated[HumanAdminRequest, Depends(require_super_admin_request)],
    connection: Annotated[Connection, Depends(get_connection)],
):
    del admin_request
    set_tenant_context(connection, tenant_id)
    _project_context(connection, tenant_id)
    content = _build_template(master_key=_DISCOUNT_POLICY_KEY)
    return StreamingResponse(
        BytesIO(content),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="mahindra-discount-policy.xlsx"'},
    )


@router.post("/segments/{segment_id}/imports", response_model=MahindraImportResponse, status_code=201)
async def upload_segment_master(
    tenant_id: str,
    segment_id: UUID,
    effective_from: Annotated[date, Form(alias="effectiveFrom")],
    file: Annotated[UploadFile, File()],
    idempotency_key: Annotated[str, Header(alias="Idempotency-Key", min_length=1)],
    admin_request: Annotated[HumanAdminRequest, Depends(require_super_admin_request)],
    connection: Annotated[Connection, Depends(get_connection)],
) -> MahindraImportResponse:
    set_tenant_context(connection, tenant_id)
    segment = _require_selected_segment(connection, tenant_id, segment_id)
    content = await file.read()
    rows = _workbook_rows(content, master_key=_SEGMENT_MASTER_KEY, segment_code=segment["segment_code"])
    validated = _validate_segment_rows(rows)
    import_id = _stage_import(
        connection,
        tenant_id=tenant_id,
        master_key=_SEGMENT_MASTER_KEY,
        segment_id=segment_id,
        effective_from=effective_from,
        file_name=file.filename or "mahindra-segment-master.xlsx",
        file_content=content,
        idempotency_key=idempotency_key,
        actor_id=admin_request.user_id,
        validated_rows=validated,
    )
    return _response(_import_record(connection, tenant_id, import_id))


@router.post("/discount-policy/imports", response_model=MahindraImportResponse, status_code=201)
async def upload_discount_policy(
    tenant_id: str,
    effective_from: Annotated[date, Form(alias="effectiveFrom")],
    file: Annotated[UploadFile, File()],
    idempotency_key: Annotated[str, Header(alias="Idempotency-Key", min_length=1)],
    admin_request: Annotated[HumanAdminRequest, Depends(require_super_admin_request)],
    connection: Annotated[Connection, Depends(get_connection)],
) -> MahindraImportResponse:
    set_tenant_context(connection, tenant_id)
    _project_context(connection, tenant_id)
    content = await file.read()
    rows = _workbook_rows(content, master_key=_DISCOUNT_POLICY_KEY)
    validated = _validate_policy_rows(connection, tenant_id, rows)
    import_id = _stage_import(
        connection,
        tenant_id=tenant_id,
        master_key=_DISCOUNT_POLICY_KEY,
        segment_id=None,
        effective_from=effective_from,
        file_name=file.filename or "mahindra-discount-policy.xlsx",
        file_content=content,
        idempotency_key=idempotency_key,
        actor_id=admin_request.user_id,
        validated_rows=validated,
    )
    return _response(_import_record(connection, tenant_id, import_id))


@router.post("/imports/{import_id}/confirm", response_model=MahindraImportResponse)
def confirm_import(
    tenant_id: str,
    import_id: UUID,
    admin_request: Annotated[HumanAdminRequest, Depends(require_super_admin_request)],
    connection: Annotated[Connection, Depends(get_connection)],
) -> MahindraImportResponse:
    set_tenant_context(connection, tenant_id)
    _project_context(connection, tenant_id)
    record = _import_record(connection, tenant_id, import_id)
    if record["status"] == "CONFIRMED":
        return _response(record, lifecycle_status="DRAFT")
    if record["status"] != "PREVIEW_READY" or int(record["error_rows"]) != 0:
        raise ConflictError(
            error_code="VAC-MASTER-003",
            title="Master import cannot be confirmed",
            detail="Only a validation-clean PREVIEW_READY import can be confirmed.",
        )

    if record["master_key"] == _SEGMENT_MASTER_KEY:
        product_version_id, price_version_id = _confirm_segment_master(
            connection,
            tenant_id=tenant_id,
            import_row=record,
            actor_id=admin_request.user_id,
        )
        receipt = {
            "productMasterVersionId": str(product_version_id),
            "priceListVersionId": str(price_version_id),
        }
        primary_version_id = product_version_id
    elif record["master_key"] == _DISCOUNT_POLICY_KEY:
        policy_version_id = _confirm_discount_policy(
            connection,
            tenant_id=tenant_id,
            import_row=record,
            actor_id=admin_request.user_id,
        )
        receipt = {"discountPolicyVersionId": str(policy_version_id)}
        primary_version_id = policy_version_id
    else:
        raise ValidationError(detail="This import is not a Mahindra master upload.")

    connection.execute(
        text(
            """
            UPDATE auditcore.project_master_imports
            SET status='CONFIRMED', confirmed_version_id=:version_id,
                confirmation_receipt=CAST(:receipt AS jsonb),
                confirmed_by_user_id=:actor_id, confirmed_at_utc=now(),
                version_no=version_no+1
            WHERE tenant_id=:tenant_id AND import_id=:import_id
            """
        ),
        {
            "version_id": primary_version_id,
            "receipt": json.dumps(receipt),
            "actor_id": admin_request.user_id,
            "tenant_id": tenant_id,
            "import_id": import_id,
        },
    )
    return _response(_import_record(connection, tenant_id, import_id), lifecycle_status="DRAFT")


@router.post("/imports/{import_id}/publish", response_model=MahindraImportResponse)
def publish_import(
    tenant_id: str,
    import_id: UUID,
    admin_request: Annotated[HumanAdminRequest, Depends(require_super_admin_request)],
    connection: Annotated[Connection, Depends(get_connection)],
) -> MahindraImportResponse:
    set_tenant_context(connection, tenant_id)
    record = _import_record(connection, tenant_id, import_id)
    if record["status"] != "CONFIRMED" or not record["confirmation_receipt"]:
        raise ConflictError(
            error_code="VAC-MASTER-003",
            title="Master import cannot be published",
            detail="Confirm the validated master import before publishing it.",
        )
    receipt = dict(record["confirmation_receipt"])
    if record["master_key"] == _SEGMENT_MASTER_KEY:
        product_version_id = UUID(str(receipt["productMasterVersionId"]))
        price_version_id = UUID(str(receipt["priceListVersionId"]))
        product_status = connection.execute(
            text(
                "SELECT lifecycle_status FROM auditcore.project_product_master_versions "
                "WHERE tenant_id=:tenant_id AND version_id=:version_id"
            ),
            {"tenant_id": tenant_id, "version_id": product_version_id},
        ).scalar_one()
        price_status = connection.execute(
            text(
                "SELECT lifecycle_status FROM auditcore.price_list_versions "
                "WHERE tenant_id=:tenant_id AND price_list_version_id=:version_id"
            ),
            {"tenant_id": tenant_id, "version_id": price_version_id},
        ).scalar_one()
        if product_status == "DRAFT":
            publish_project_product_master_version(
                connection,
                tenant_id=tenant_id,
                version_id=product_version_id,
                actor_id=admin_request.user_id,
            )
        if price_status == "DRAFT":
            publish_price_list_version(
                connection,
                tenant_id=tenant_id,
                price_list_version_id=price_version_id,
                actor_id=admin_request.user_id,
            )
    elif record["master_key"] == _DISCOUNT_POLICY_KEY:
        version_id = UUID(str(receipt["discountPolicyVersionId"]))
        connection.execute(
            text(
                """
                UPDATE auditcore.discount_policy_versions
                SET lifecycle_status='PUBLISHED', published_by_actor_id=:actor_id,
                    published_at_utc=now(), updated_at_utc=now()
                WHERE tenant_id=:tenant_id
                  AND discount_policy_version_id=:version_id
                  AND lifecycle_status='DRAFT'
                """
            ),
            {"tenant_id": tenant_id, "version_id": version_id, "actor_id": admin_request.user_id},
        )
    else:
        raise ValidationError(detail="This import is not a Mahindra master upload.")
    return _response(_import_record(connection, tenant_id, import_id), lifecycle_status="PUBLISHED")
