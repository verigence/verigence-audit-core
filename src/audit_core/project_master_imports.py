from __future__ import annotations

import hashlib
import json
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import Annotated, Any, Literal
from uuid import UUID

from fastapi import APIRouter, Depends, File, Form, Header, Query, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from pydantic import BaseModel
from sqlalchemy import Connection, text

from audit_core.db import set_tenant_context
from audit_core.dependencies import (
    HumanAdminRequest,
    get_connection,
    require_super_admin_request,
)
from audit_core.discount_schemes import (
    add_discount_benefit,
    add_discount_eligibility,
    create_discount_scheme,
    create_discount_scheme_version,
)
from audit_core.errors import AuditCoreError, ConflictError, NotFoundError, ValidationError
from audit_core.idempotency import stable_request_hash
from audit_core.price_lists import add_price_list_item, create_price_list, create_price_list_version
from audit_core.product_masters import (
    add_project_product_master_item,
    create_project_product_master_version,
    product_sku_is_in_effective_master,
)
from audit_core.project_masters import descriptor_definition, excel_master_keys

router = APIRouter(tags=["project-master-imports"])

ValidationStatus = Literal["VALID", "WARNING", "ERROR"]
ImportStatus = Literal[
    "UPLOADED",
    "PARSING",
    "PREVIEW_READY",
    "VALIDATION_FAILED",
    "CONFIRMED",
    "CANCELLED",
    "FAILED",
]


class ProjectMasterImportResponse(BaseModel):
    importId: UUID
    ownerModule: str
    masterKey: str
    effectiveFrom: date | None
    templateVersion: str | None
    fileName: str
    fileHash: str
    status: ImportStatus
    rowsParsed: int
    validRows: int
    warningRows: int
    errorRows: int
    confirmedVersionId: UUID | None
    createdBy: str
    createdAtUtc: datetime
    confirmedBy: str | None
    confirmedAtUtc: datetime | None
    versionNo: int


class ImportRowResponse(BaseModel):
    rowNumber: int
    parsedData: dict[str, Any]
    validationStatus: ValidationStatus
    messages: list[str]


class ImportRowsPage(BaseModel):
    items: list[ImportRowResponse]
    offset: int
    limit: int
    total: int


_TEMPLATE_COLUMNS: dict[str, tuple[str, ...]] = {
    "PRODUCT_MASTER": (
        "sku_code",
        "model_code",
        "model_name",
        "model_year",
        "variant_code",
        "variant_name",
        "fuel_powertrain",
        "transmission",
        "body_type",
        "colour_code",
        "colour_name",
        "sku_attributes_json",
        "variant_attributes_json",
    ),
    "PRICE_LIST": (
        "price_list_code",
        "price_list_name",
        "sku_code",
        "component_key",
        "standard_amount",
        "currency_code",
        "effective_to",
    ),
    "DISCOUNT_SCHEME": (
        "scheme_code",
        "scheme_name",
        "scheme_category",
        "sku_code",
        "dealer_id",
        "outlet_id",
        "customer_type_code",
        "benefit_key",
        "benefit_type",
        "amount_value",
        "percentage_value",
        "effective_to",
    ),
}

_REQUIRED_COLUMNS: dict[str, frozenset[str]] = {
    "PRODUCT_MASTER": frozenset(
        {"sku_code", "model_code", "model_name", "variant_code", "variant_name"}
    ),
    "PRICE_LIST": frozenset(
        {
            "price_list_code",
            "price_list_name",
            "sku_code",
            "component_key",
            "standard_amount",
        }
    ),
    "DISCOUNT_SCHEME": frozenset(
        {"scheme_code", "scheme_name", "benefit_key", "benefit_type"}
    ),
}


def _not_found() -> NotFoundError:
    return NotFoundError(
        error_code="VAC-NF-008",
        title="Project Master import not found",
        detail="The requested Project Master import does not exist for this Project.",
    )


def _import_row(connection: Connection, *, tenant_id: str, import_id: UUID):
    row = connection.execute(
        text(
            """
            SELECT import_id, owner_module, master_key, effective_from,
                   template_version, original_file_name, file_hash, status,
                   rows_parsed, valid_rows, warning_rows, error_rows,
                   confirmed_version_id, created_by_user_id, created_at_utc,
                   confirmed_by_user_id, confirmed_at_utc, version_no,
                   idempotency_key, semantic_request_hash
            FROM auditcore.project_master_imports
            WHERE tenant_id=:tenant_id AND import_id=:import_id
            """
        ),
        {"tenant_id": tenant_id, "import_id": import_id},
    ).mappings().one_or_none()
    if row is None:
        raise _not_found()
    return row


def _response(row) -> ProjectMasterImportResponse:
    return ProjectMasterImportResponse(
        importId=row["import_id"],
        ownerModule=row["owner_module"],
        masterKey=row["master_key"],
        effectiveFrom=row["effective_from"],
        templateVersion=row["template_version"],
        fileName=row["original_file_name"],
        fileHash=row["file_hash"],
        status=row["status"],
        rowsParsed=int(row["rows_parsed"]),
        validRows=int(row["valid_rows"]),
        warningRows=int(row["warning_rows"]),
        errorRows=int(row["error_rows"]),
        confirmedVersionId=row["confirmed_version_id"],
        createdBy=row["created_by_user_id"],
        createdAtUtc=row["created_at_utc"],
        confirmedBy=row["confirmed_by_user_id"],
        confirmedAtUtc=row["confirmed_at_utc"],
        versionNo=int(row["version_no"]),
    )


def _template_version(master_key: str) -> str:
    descriptor = descriptor_definition(master_key)
    value = descriptor["templateVersion"]
    if value is None:
        raise AuditCoreError(
            error_code="VAC-MASTER-004",
            status_code=422,
            title="Excel administration unsupported",
            detail="The requested Project Master is not registered for Excel administration.",
        )
    return str(value)


def build_template(master_key: str) -> bytes:
    normalized = master_key.strip().upper()
    if normalized not in excel_master_keys():
        descriptor_definition(normalized)
        raise AuditCoreError(
            error_code="VAC-MASTER-004",
            status_code=422,
            title="Excel administration unsupported",
            detail="The requested Project Master is not registered for Excel administration.",
        )
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = normalized[:31]
    for index, column in enumerate(_TEMPLATE_COLUMNS[normalized], start=1):
        sheet.cell(row=1, column=index, value=column)
    metadata = workbook.create_sheet("_meta")
    metadata["A1"] = "master_key"
    metadata["B1"] = normalized
    metadata["A2"] = "template_version"
    metadata["B2"] = _template_version(normalized)
    metadata.sheet_state = "hidden"
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _normalise_cell(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, Decimal):
        return str(value)
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return value


def _workbook_rows(content: bytes, *, master_key: str) -> list[tuple[int, dict[str, Any]]]:
    try:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise ValidationError(detail="The uploaded .xlsx workbook is unreadable.") from exc
    if "_meta" in workbook.sheetnames:
        metadata = workbook["_meta"]
        workbook_master_key = metadata["B1"].value
        workbook_template_version = metadata["B2"].value
        if workbook_master_key != master_key:
            raise ValidationError(detail="Workbook master metadata does not match the selected master.")
        if str(workbook_template_version) != _template_version(master_key):
            raise ValidationError(detail="Workbook template version is not supported.")

    data_sheets = [name for name in workbook.sheetnames if name != "_meta"]
    if len(data_sheets) != 1:
        raise ValidationError(detail="Project Master workbook must contain exactly one data sheet.")
    sheet = workbook[data_sheets[0]]
    iterator = sheet.iter_rows(values_only=True)
    try:
        raw_headers = next(iterator)
    except StopIteration as exc:
        raise ValidationError(detail="Project Master workbook is empty.") from exc
    headers = [str(value).strip() if value is not None else "" for value in raw_headers]
    if not headers or any(not header for header in headers):
        raise ValidationError(detail="Project Master workbook header contains an empty column name.")
    if len(headers) != len(set(headers)):
        raise ValidationError(detail="Project Master workbook header contains duplicate columns.")
    allowed = set(_TEMPLATE_COLUMNS[master_key])
    unknown = [header for header in headers if header not in allowed]
    missing = sorted(_REQUIRED_COLUMNS[master_key] - set(headers))
    if unknown:
        raise ValidationError(
            detail=f"Workbook contains unsupported column(s): {', '.join(sorted(unknown))}."
        )
    if missing:
        raise ValidationError(
            detail=f"Workbook is missing required column(s): {', '.join(missing)}."
        )

    rows: list[tuple[int, dict[str, Any]]] = []
    for row_number, values in enumerate(iterator, start=2):
        if all(value is None or str(value).strip() == "" for value in values):
            continue
        parsed = {
            header: _normalise_cell(value)
            for header, value in zip(headers, values, strict=False)
        }
        rows.append((row_number, parsed))
    if not rows:
        raise ValidationError(detail="Project Master workbook contains no data rows.")
    return rows


def _required_text(row: dict[str, Any], key: str, messages: list[str]) -> str | None:
    value = row.get(key)
    if value is None or not str(value).strip():
        messages.append(f"{key} is required.")
        return None
    return str(value).strip()


def _optional_text(row: dict[str, Any], key: str) -> str | None:
    value = row.get(key)
    if value is None or not str(value).strip():
        return None
    return str(value).strip()


def _decimal_value(
    row: dict[str, Any],
    key: str,
    messages: list[str],
    *,
    required: bool = False,
) -> Decimal | None:
    value = row.get(key)
    if value is None or str(value).strip() == "":
        if required:
            messages.append(f"{key} is required.")
        return None
    try:
        return Decimal(str(value))
    except InvalidOperation:
        messages.append(f"{key} must be numeric.")
        return None


def _date_value(row: dict[str, Any], key: str, messages: list[str]) -> date | None:
    value = row.get(key)
    if value is None or str(value).strip() == "":
        return None
    if isinstance(value, date):
        return value
    try:
        return date.fromisoformat(str(value).strip())
    except ValueError:
        messages.append(f"{key} must be an ISO date (YYYY-MM-DD).")
        return None


def _json_object(row: dict[str, Any], key: str, messages: list[str]) -> dict[str, Any]:
    value = row.get(key)
    if value is None or str(value).strip() == "":
        return {}
    if isinstance(value, dict):
        return value
    try:
        parsed = json.loads(str(value))
    except json.JSONDecodeError:
        messages.append(f"{key} must contain a JSON object.")
        return {}
    if not isinstance(parsed, dict):
        messages.append(f"{key} must contain a JSON object.")
        return {}
    return parsed


def _project_oem_id(connection: Connection, tenant_id: str) -> UUID:
    value = connection.execute(
        text("SELECT oem_id FROM auditcore.projects WHERE tenant_id=:tenant_id"),
        {"tenant_id": tenant_id},
    ).scalar_one_or_none()
    if value is None:
        raise NotFoundError(
            error_code="VAC-NF-001",
            title="Project not found",
            detail="Project not found for the requested tenant.",
        )
    return value


def _sku_by_code(connection: Connection, *, tenant_id: str, sku_code: str):
    oem_id = _project_oem_id(connection, tenant_id)
    return connection.execute(
        text(
            """
            SELECT s.product_sku_id, s.model_id, s.variant_id, s.colour_id,
                   m.model_code, v.variant_code, c.colour_code
            FROM auditcore.product_skus s
            JOIN auditcore.product_models m ON m.model_id=s.model_id
            JOIN auditcore.product_variants v ON v.variant_id=s.variant_id
            LEFT JOIN auditcore.colours c ON c.colour_id=s.colour_id
            WHERE s.oem_id=:oem_id AND s.sku_code=:sku_code
            """
        ),
        {"oem_id": oem_id, "sku_code": sku_code},
    ).mappings().one_or_none()


def _effective_sku(
    connection: Connection,
    *,
    tenant_id: str,
    sku_code: str,
    effective_from: date,
    messages: list[str],
) -> UUID | None:
    sku = _sku_by_code(connection, tenant_id=tenant_id, sku_code=sku_code)
    if sku is None:
        messages.append(f"sku_code {sku_code!r} does not reference a canonical Project OEM SKU.")
        return None
    try:
        in_master = product_sku_is_in_effective_master(
            connection,
            tenant_id=tenant_id,
            product_sku_id=sku["product_sku_id"],
            effective_on=effective_from,
        )
    except AuditCoreError:
        messages.append(
            "No unambiguous published Product Master is effective for this import WEF."
        )
        return None
    if not in_master:
        messages.append(
            f"sku_code {sku_code!r} is not present in the Product Master effective for this WEF."
        )
        return None
    return sku["product_sku_id"]


def _validate_product_row(
    connection: Connection,
    *,
    tenant_id: str,
    row: dict[str, Any],
) -> list[str]:
    messages: list[str] = []
    sku_code = _required_text(row, "sku_code", messages)
    model_code = _required_text(row, "model_code", messages)
    _required_text(row, "model_name", messages)
    variant_code = _required_text(row, "variant_code", messages)
    _required_text(row, "variant_name", messages)
    colour_code = _optional_text(row, "colour_code")
    colour_name = _optional_text(row, "colour_name")
    if colour_code and not colour_name:
        messages.append("colour_name is required when colour_code is supplied.")
    if colour_name and not colour_code:
        messages.append("colour_code is required when colour_name is supplied.")
    model_year = row.get("model_year")
    if model_year not in (None, ""):
        try:
            year = int(model_year)
            if year < 1900 or year > 2200:
                raise ValueError
        except (TypeError, ValueError):
            messages.append("model_year must be a four-digit year.")
    _json_object(row, "sku_attributes_json", messages)
    _json_object(row, "variant_attributes_json", messages)

    if sku_code and model_code and variant_code:
        existing = _sku_by_code(connection, tenant_id=tenant_id, sku_code=sku_code)
        if existing is not None:
            if existing["model_code"] != model_code:
                messages.append("Existing sku_code is assigned to a different model_code.")
            if existing["variant_code"] != variant_code:
                messages.append("Existing sku_code is assigned to a different variant_code.")
            existing_colour = existing["colour_code"]
            if (existing_colour or None) != (colour_code or None):
                messages.append("Existing sku_code is assigned to a different colour_code.")
    return messages


def _validate_price_row(
    connection: Connection,
    *,
    tenant_id: str,
    effective_from: date,
    row: dict[str, Any],
) -> list[str]:
    messages: list[str] = []
    _required_text(row, "price_list_code", messages)
    _required_text(row, "price_list_name", messages)
    sku_code = _required_text(row, "sku_code", messages)
    _required_text(row, "component_key", messages)
    amount = _decimal_value(row, "standard_amount", messages, required=True)
    if amount is not None and amount < 0:
        messages.append("standard_amount cannot be negative.")
    currency = _optional_text(row, "currency_code")
    if currency is not None and (len(currency) != 3 or not currency.isalpha()):
        messages.append("currency_code must be a three-letter currency code.")
    effective_to = _date_value(row, "effective_to", messages)
    if effective_to is not None and effective_to < effective_from:
        messages.append("effective_to cannot be earlier than the import WEF.")
    if sku_code:
        _effective_sku(
            connection,
            tenant_id=tenant_id,
            sku_code=sku_code,
            effective_from=effective_from,
            messages=messages,
        )
    return messages


def _validate_discount_row(
    connection: Connection,
    *,
    tenant_id: str,
    effective_from: date,
    row: dict[str, Any],
) -> list[str]:
    messages: list[str] = []
    _required_text(row, "scheme_code", messages)
    _required_text(row, "scheme_name", messages)
    _required_text(row, "benefit_key", messages)
    benefit_type = _required_text(row, "benefit_type", messages)
    if benefit_type and benefit_type.upper() not in {"AMOUNT", "PERCENTAGE", "OTHER"}:
        messages.append("benefit_type must be AMOUNT, PERCENTAGE or OTHER.")
    amount = _decimal_value(row, "amount_value", messages)
    percentage = _decimal_value(row, "percentage_value", messages)
    if benefit_type:
        upper = benefit_type.upper()
        if upper == "AMOUNT" and amount is None:
            messages.append("amount_value is required for AMOUNT benefits.")
        if upper == "PERCENTAGE" and percentage is None:
            messages.append("percentage_value is required for PERCENTAGE benefits.")
    if amount is not None and amount < 0:
        messages.append("amount_value cannot be negative.")
    if percentage is not None and percentage < 0:
        messages.append("percentage_value cannot be negative.")

    sku_code = _optional_text(row, "sku_code")
    if sku_code:
        _effective_sku(
            connection,
            tenant_id=tenant_id,
            sku_code=sku_code,
            effective_from=effective_from,
            messages=messages,
        )

    dealer_id = _optional_text(row, "dealer_id")
    outlet_id = _optional_text(row, "outlet_id")
    if outlet_id and not dealer_id:
        messages.append("dealer_id is required when outlet_id is supplied.")
    if dealer_id:
        dealer_exists = connection.execute(
            text(
                "SELECT 1 FROM auditcore.dealers "
                "WHERE tenant_id=:tenant_id AND dealer_id=CAST(:dealer_id AS uuid)"
            ),
            {"tenant_id": tenant_id, "dealer_id": dealer_id},
        ).scalar_one_or_none()
        if dealer_exists is None:
            messages.append("dealer_id does not belong to this Project.")
    if outlet_id and dealer_id:
        outlet_exists = connection.execute(
            text(
                """
                SELECT 1 FROM auditcore.dealer_outlets
                WHERE tenant_id=:tenant_id
                  AND dealer_id=CAST(:dealer_id AS uuid)
                  AND outlet_id=CAST(:outlet_id AS uuid)
                """
            ),
            {"tenant_id": tenant_id, "dealer_id": dealer_id, "outlet_id": outlet_id},
        ).scalar_one_or_none()
        if outlet_exists is None:
            messages.append("outlet_id does not belong to the selected Project Dealer.")

    effective_to = _date_value(row, "effective_to", messages)
    if effective_to is not None and effective_to < effective_from:
        messages.append("effective_to cannot be earlier than the import WEF.")
    return messages


def _cross_row_messages(
    master_key: str,
    rows: list[tuple[int, dict[str, Any]]],
) -> dict[int, list[str]]:
    messages: dict[int, list[str]] = {row_number: [] for row_number, _ in rows}
    if master_key == "PRODUCT_MASTER":
        seen: dict[str, int] = {}
        for row_number, row in rows:
            sku_code = _optional_text(row, "sku_code")
            if not sku_code:
                continue
            if sku_code in seen:
                messages[row_number].append(
                    f"sku_code duplicates workbook row {seen[sku_code]}."
                )
            else:
                seen[sku_code] = row_number
    elif master_key == "PRICE_LIST":
        identity = None
        for row_number, row in rows:
            current = (
                _optional_text(row, "price_list_code"),
                _optional_text(row, "price_list_name"),
                (_optional_text(row, "currency_code") or "INR").upper(),
                _optional_text(row, "effective_to"),
            )
            if identity is None:
                identity = current
            elif current != identity:
                messages[row_number].append(
                    "All rows in one Price List import must use the same list identity, "
                    "currency and effective_to."
                )
    elif master_key == "DISCOUNT_SCHEME":
        identity = None
        seen_benefits: dict[str, int] = {}
        for row_number, row in rows:
            current = (
                _optional_text(row, "scheme_code"),
                _optional_text(row, "scheme_name"),
                _optional_text(row, "scheme_category"),
                _optional_text(row, "effective_to"),
            )
            if identity is None:
                identity = current
            elif current != identity:
                messages[row_number].append(
                    "All rows in one Discount Scheme import must use the same scheme "
                    "identity and effective_to."
                )
            benefit_key = _optional_text(row, "benefit_key")
            if benefit_key:
                if benefit_key in seen_benefits:
                    messages[row_number].append(
                        f"benefit_key duplicates workbook row {seen_benefits[benefit_key]}."
                    )
                else:
                    seen_benefits[benefit_key] = row_number
    return messages


def _overlap_warning(
    connection: Connection,
    *,
    tenant_id: str,
    master_key: str,
    effective_from: date,
    effective_to: date | None,
) -> str | None:
    if master_key == "PRODUCT_MASTER":
        exists = connection.execute(
            text(
                """
                SELECT EXISTS (
                    SELECT 1 FROM auditcore.project_product_master_versions
                    WHERE tenant_id=:tenant_id
                      AND lifecycle_status='PUBLISHED'
                      AND effective_from <= :effective_from
                )
                """
            ),
            {"tenant_id": tenant_id, "effective_from": effective_from},
        ).scalar_one()
    else:
        table, start_column, end_column = {
            "PRICE_LIST": ("price_list_versions", "effective_from", "effective_to"),
            "DISCOUNT_SCHEME": (
                "discount_scheme_versions",
                "effective_from",
                "effective_to",
            ),
        }[master_key]
        exists = connection.execute(
            text(
                f"""
                SELECT EXISTS (
                    SELECT 1 FROM auditcore.{table}
                    WHERE tenant_id=:tenant_id
                      AND lifecycle_status='PUBLISHED'
                      AND ({end_column} IS NULL OR {end_column} >= :effective_from)
                      AND (CAST(:effective_to AS date) IS NULL
                           OR {start_column} <= CAST(:effective_to AS date))
                )
                """
            ),
            {
                "tenant_id": tenant_id,
                "effective_from": effective_from,
                "effective_to": effective_to,
            },
        ).scalar_one()
    if exists:
        return (
            "The import WEF overlaps an existing published version. Phase 1 permits overlap; "
            "the owning master resolver determines the effective version."
        )
    return None


def _validate_rows(
    connection: Connection,
    *,
    tenant_id: str,
    master_key: str,
    effective_from: date,
    rows: list[tuple[int, dict[str, Any]]],
) -> list[tuple[int, dict[str, Any], ValidationStatus, list[str]]]:
    cross_messages = _cross_row_messages(master_key, rows)
    validated: list[tuple[int, dict[str, Any], ValidationStatus, list[str]]] = []
    for row_number, row in rows:
        if master_key == "PRODUCT_MASTER":
            messages = _validate_product_row(connection, tenant_id=tenant_id, row=row)
        elif master_key == "PRICE_LIST":
            messages = _validate_price_row(
                connection,
                tenant_id=tenant_id,
                effective_from=effective_from,
                row=row,
            )
        else:
            messages = _validate_discount_row(
                connection,
                tenant_id=tenant_id,
                effective_from=effective_from,
                row=row,
            )
        messages.extend(cross_messages[row_number])
        effective_to_messages: list[str] = []
        effective_to = _date_value(row, "effective_to", effective_to_messages)
        messages.extend(effective_to_messages)
        warning = _overlap_warning(
            connection,
            tenant_id=tenant_id,
            master_key=master_key,
            effective_from=effective_from,
            effective_to=effective_to,
        )
        status: ValidationStatus = "ERROR" if messages else "VALID"
        if warning and status == "VALID":
            status = "WARNING"
            messages = [warning]
        validated.append((row_number, row, status, messages))
    return validated


def _insert_validated_rows(
    connection: Connection,
    *,
    tenant_id: str,
    import_id: UUID,
    rows: list[tuple[int, dict[str, Any], ValidationStatus, list[str]]],
) -> tuple[int, int, int]:
    valid = warning = error = 0
    for row_number, parsed, status, messages in rows:
        valid += status == "VALID"
        warning += status == "WARNING"
        error += status == "ERROR"
        connection.execute(
            text(
                """
                INSERT INTO auditcore.project_master_import_rows (
                    tenant_id, import_id, row_number, parsed_data,
                    validation_status, validation_messages
                ) VALUES (
                    :tenant_id, :import_id, :row_number, CAST(:parsed_data AS jsonb),
                    :validation_status, CAST(:messages AS jsonb)
                )
                """
            ),
            {
                "tenant_id": tenant_id,
                "import_id": import_id,
                "row_number": row_number,
                "parsed_data": json.dumps(parsed, default=str),
                "validation_status": status,
                "messages": json.dumps(messages),
            },
        )
    return int(valid), int(warning), int(error)


def _load_parsed_rows(
    connection: Connection,
    *,
    tenant_id: str,
    import_id: UUID,
) -> list[dict[str, Any]]:
    return [
        dict(row["parsed_data"])
        for row in connection.execute(
            text(
                """
                SELECT parsed_data
                FROM auditcore.project_master_import_rows
                WHERE tenant_id=:tenant_id AND import_id=:import_id
                ORDER BY row_number
                """
            ),
            {"tenant_id": tenant_id, "import_id": import_id},
        ).mappings()
    ]


def _resolve_or_create_product_sku(
    connection: Connection,
    *,
    tenant_id: str,
    row: dict[str, Any],
) -> UUID:
    oem_id = _project_oem_id(connection, tenant_id)
    model_code = str(row["model_code"]).strip()
    model_name = str(row["model_name"]).strip()
    model = connection.execute(
        text(
            "SELECT model_id, model_name FROM auditcore.product_models "
            "WHERE oem_id=:oem_id AND model_code=:model_code"
        ),
        {"oem_id": oem_id, "model_code": model_code},
    ).mappings().one_or_none()
    if model is None:
        model_id = connection.execute(
            text(
                """
                INSERT INTO auditcore.product_models (
                    oem_id, model_code, model_name, model_year
                ) VALUES (
                    :oem_id, :model_code, :model_name, :model_year
                ) RETURNING model_id
                """
            ),
            {
                "oem_id": oem_id,
                "model_code": model_code,
                "model_name": model_name,
                "model_year": int(row["model_year"]) if row.get("model_year") not in (None, "") else None,
            },
        ).scalar_one()
    else:
        if model["model_name"] != model_name:
            raise ConflictError(
                error_code="VAC-CONFLICT-004",
                title="Canonical Product conflict",
                detail="Existing model_code has a different canonical model name.",
            )
        model_id = model["model_id"]

    variant_code = str(row["variant_code"]).strip()
    variant_name = str(row["variant_name"]).strip()
    variant = connection.execute(
        text(
            "SELECT variant_id, variant_name FROM auditcore.product_variants "
            "WHERE model_id=:model_id AND variant_code=:variant_code"
        ),
        {"model_id": model_id, "variant_code": variant_code},
    ).mappings().one_or_none()
    if variant is None:
        variant_id = connection.execute(
            text(
                """
                INSERT INTO auditcore.product_variants (
                    model_id, variant_code, variant_name, fuel_powertrain,
                    transmission, body_type, attributes
                ) VALUES (
                    :model_id, :variant_code, :variant_name, :fuel_powertrain,
                    :transmission, :body_type, CAST(:attributes AS jsonb)
                ) RETURNING variant_id
                """
            ),
            {
                "model_id": model_id,
                "variant_code": variant_code,
                "variant_name": variant_name,
                "fuel_powertrain": _optional_text(row, "fuel_powertrain"),
                "transmission": _optional_text(row, "transmission"),
                "body_type": _optional_text(row, "body_type"),
                "attributes": json.dumps(
                    _json_object(row, "variant_attributes_json", [])
                ),
            },
        ).scalar_one()
    else:
        if variant["variant_name"] != variant_name:
            raise ConflictError(
                error_code="VAC-CONFLICT-004",
                title="Canonical Product conflict",
                detail="Existing variant_code has a different canonical variant name.",
            )
        variant_id = variant["variant_id"]

    colour_id = None
    colour_code = _optional_text(row, "colour_code")
    if colour_code:
        colour_name = str(row["colour_name"]).strip()
        colour = connection.execute(
            text(
                "SELECT colour_id, colour_name FROM auditcore.colours "
                "WHERE oem_id=:oem_id AND colour_code=:colour_code"
            ),
            {"oem_id": oem_id, "colour_code": colour_code},
        ).mappings().one_or_none()
        if colour is None:
            colour_id = connection.execute(
                text(
                    """
                    INSERT INTO auditcore.colours (oem_id, colour_code, colour_name)
                    VALUES (:oem_id, :colour_code, :colour_name)
                    RETURNING colour_id
                    """
                ),
                {
                    "oem_id": oem_id,
                    "colour_code": colour_code,
                    "colour_name": colour_name,
                },
            ).scalar_one()
        else:
            if colour["colour_name"] != colour_name:
                raise ConflictError(
                    error_code="VAC-CONFLICT-004",
                    title="Canonical Product conflict",
                    detail="Existing colour_code has a different canonical colour name.",
                )
            colour_id = colour["colour_id"]

    sku_code = str(row["sku_code"]).strip()
    sku = _sku_by_code(connection, tenant_id=tenant_id, sku_code=sku_code)
    if sku is not None:
        if (
            sku["model_id"] != model_id
            or sku["variant_id"] != variant_id
            or sku["colour_id"] != colour_id
        ):
            raise ConflictError(
                error_code="VAC-CONFLICT-004",
                title="Canonical Product conflict",
                detail="Existing sku_code points to a different canonical Product configuration.",
            )
        return sku["product_sku_id"]
    return connection.execute(
        text(
            """
            INSERT INTO auditcore.product_skus (
                oem_id, model_id, variant_id, colour_id, sku_code, attributes
            ) VALUES (
                :oem_id, :model_id, :variant_id, :colour_id, :sku_code,
                CAST(:attributes AS jsonb)
            ) RETURNING product_sku_id
            """
        ),
        {
            "oem_id": oem_id,
            "model_id": model_id,
            "variant_id": variant_id,
            "colour_id": colour_id,
            "sku_code": sku_code,
            "attributes": json.dumps(_json_object(row, "sku_attributes_json", [])),
        },
    ).scalar_one()


def _confirm_product(
    connection: Connection,
    *,
    tenant_id: str,
    effective_from: date,
    actor_id: str,
    import_id: UUID,
    rows: list[dict[str, Any]],
) -> UUID:
    version_id = create_project_product_master_version(
        connection,
        tenant_id=tenant_id,
        effective_from=effective_from,
        actor_id=actor_id,
        source_import_id=import_id,
    )
    for index, row in enumerate(rows, start=2):
        sku_id = _resolve_or_create_product_sku(
            connection,
            tenant_id=tenant_id,
            row=row,
        )
        add_project_product_master_item(
            connection,
            tenant_id=tenant_id,
            version_id=version_id,
            product_sku_id=sku_id,
            source_import_row_no=index,
        )
    return version_id


def _confirm_price(
    connection: Connection,
    *,
    tenant_id: str,
    effective_from: date,
    actor_id: str,
    rows: list[dict[str, Any]],
) -> UUID:
    first = rows[0]
    code = str(first["price_list_code"]).strip()
    name = str(first["price_list_name"]).strip()
    price_list = connection.execute(
        text(
            "SELECT price_list_id, price_list_name FROM auditcore.price_lists "
            "WHERE tenant_id=:tenant_id AND price_list_code=:code"
        ),
        {"tenant_id": tenant_id, "code": code},
    ).mappings().one_or_none()
    if price_list is None:
        price_list_id = create_price_list(
            connection,
            tenant_id=tenant_id,
            code=code,
            name=name,
            actor_id=actor_id,
        )
    else:
        if price_list["price_list_name"] != name:
            raise ConflictError(
                error_code="VAC-CONFLICT-005",
                title="Price List identity conflict",
                detail="Existing Price List code has a different name.",
            )
        price_list_id = price_list["price_list_id"]
    version_no = connection.execute(
        text(
            "SELECT COALESCE(MAX(version_no),0)+1 FROM auditcore.price_list_versions "
            "WHERE tenant_id=:tenant_id AND price_list_id=:price_list_id"
        ),
        {"tenant_id": tenant_id, "price_list_id": price_list_id},
    ).scalar_one()
    effective_to = _date_value(first, "effective_to", [])
    currency = (_optional_text(first, "currency_code") or "INR").upper()
    version_id = create_price_list_version(
        connection,
        tenant_id=tenant_id,
        price_list_id=price_list_id,
        version_no=version_no,
        effective_from=effective_from,
        effective_to=effective_to,
        currency_code=currency,
        actor_id=actor_id,
    )
    for row in rows:
        sku = _sku_by_code(
            connection,
            tenant_id=tenant_id,
            sku_code=str(row["sku_code"]).strip(),
        )
        if sku is None:
            raise ConflictError(
                error_code="VAC-CONFLICT-006",
                title="Product reference conflict",
                detail="Validated Price List SKU cannot be resolved at confirmation time.",
            )
        add_price_list_item(
            connection,
            tenant_id=tenant_id,
            price_list_version_id=version_id,
            product_sku_id=sku["product_sku_id"],
            component_key=str(row["component_key"]).strip(),
            standard_amount=Decimal(str(row["standard_amount"])),
        )
    return version_id


def _confirm_discount(
    connection: Connection,
    *,
    tenant_id: str,
    effective_from: date,
    actor_id: str,
    rows: list[dict[str, Any]],
) -> UUID:
    first = rows[0]
    code = str(first["scheme_code"]).strip()
    name = str(first["scheme_name"]).strip()
    category = _optional_text(first, "scheme_category")
    scheme = connection.execute(
        text(
            "SELECT discount_scheme_id, scheme_name, scheme_category "
            "FROM auditcore.discount_schemes "
            "WHERE tenant_id=:tenant_id AND scheme_code=:code"
        ),
        {"tenant_id": tenant_id, "code": code},
    ).mappings().one_or_none()
    if scheme is None:
        scheme_id = create_discount_scheme(
            connection,
            tenant_id=tenant_id,
            code=code,
            name=name,
            category=category,
            actor_id=actor_id,
        )
    else:
        if scheme["scheme_name"] != name or scheme["scheme_category"] != category:
            raise ConflictError(
                error_code="VAC-CONFLICT-007",
                title="Discount Scheme identity conflict",
                detail="Existing Discount Scheme code has different identity metadata.",
            )
        scheme_id = scheme["discount_scheme_id"]
    version_no = connection.execute(
        text(
            "SELECT COALESCE(MAX(version_no),0)+1 FROM auditcore.discount_scheme_versions "
            "WHERE tenant_id=:tenant_id AND discount_scheme_id=:scheme_id"
        ),
        {"tenant_id": tenant_id, "scheme_id": scheme_id},
    ).scalar_one()
    version_id = create_discount_scheme_version(
        connection,
        tenant_id=tenant_id,
        discount_scheme_id=scheme_id,
        version_no=version_no,
        effective_from=effective_from,
        effective_to=_date_value(first, "effective_to", []),
        actor_id=actor_id,
    )
    for row in rows:
        sku_id = None
        sku_code = _optional_text(row, "sku_code")
        if sku_code:
            sku = _sku_by_code(connection, tenant_id=tenant_id, sku_code=sku_code)
            if sku is None:
                raise ConflictError(
                    error_code="VAC-CONFLICT-006",
                    title="Product reference conflict",
                    detail="Validated Discount Scheme SKU cannot be resolved at confirmation time.",
                )
            sku_id = sku["product_sku_id"]
        dealer_id = UUID(value) if (value := _optional_text(row, "dealer_id")) else None
        outlet_id = UUID(value) if (value := _optional_text(row, "outlet_id")) else None
        if any(
            value is not None
            for value in (
                sku_id,
                dealer_id,
                outlet_id,
                _optional_text(row, "customer_type_code"),
            )
        ):
            add_discount_eligibility(
                connection,
                tenant_id=tenant_id,
                discount_scheme_version_id=version_id,
                product_sku_id=sku_id,
                dealer_id=dealer_id,
                outlet_id=outlet_id,
                customer_type_code=_optional_text(row, "customer_type_code"),
            )
        add_discount_benefit(
            connection,
            tenant_id=tenant_id,
            discount_scheme_version_id=version_id,
            benefit_key=str(row["benefit_key"]).strip(),
            benefit_type=str(row["benefit_type"]).strip().upper(),
            amount_value=(
                Decimal(str(row["amount_value"]))
                if row.get("amount_value") not in (None, "")
                else None
            ),
            percentage_value=(
                Decimal(str(row["percentage_value"]))
                if row.get("percentage_value") not in (None, "")
                else None
            ),
        )
    return version_id


def _confirm_import(
    connection: Connection,
    *,
    tenant_id: str,
    master_key: str,
    effective_from: date,
    actor_id: str,
    import_id: UUID,
    rows: list[dict[str, Any]],
) -> UUID:
    if master_key == "PRODUCT_MASTER":
        return _confirm_product(
            connection,
            tenant_id=tenant_id,
            effective_from=effective_from,
            actor_id=actor_id,
            import_id=import_id,
            rows=rows,
        )
    if master_key == "PRICE_LIST":
        return _confirm_price(
            connection,
            tenant_id=tenant_id,
            effective_from=effective_from,
            actor_id=actor_id,
            rows=rows,
        )
    if master_key == "DISCOUNT_SCHEME":
        return _confirm_discount(
            connection,
            tenant_id=tenant_id,
            effective_from=effective_from,
            actor_id=actor_id,
            rows=rows,
        )
    raise AuditCoreError(
        error_code="VAC-MASTER-004",
        status_code=422,
        title="Excel administration unsupported",
        detail="The requested Project Master is not registered for Excel confirmation.",
    )


@router.get(
    "/v1/tenants/{tenant_id}/project-masters/{owner_module}/{master_key}/template"
)
def download_master_template(
    tenant_id: str,
    owner_module: str,
    master_key: str,
    admin_request: Annotated[HumanAdminRequest, Depends(require_super_admin_request)],
    connection: Annotated[Connection, Depends(get_connection)],
) -> StreamingResponse:
    set_tenant_context(connection, tenant_id)
    if owner_module != "AUDIT_CORE":
        raise AuditCoreError(
            error_code="VAC-MASTER-004",
            status_code=422,
            title="Owning module required",
            detail="DI-owned templates are supplied by DI through the Audit Core facade.",
        )
    normalized = master_key.strip().upper()
    content = build_template(normalized)
    return StreamingResponse(
        BytesIO(content),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{normalized.lower()}-template.xlsx"'
        },
    )


@router.post(
    "/v1/tenants/{tenant_id}/project-masters/{owner_module}/{master_key}/imports",
    response_model=ProjectMasterImportResponse,
    status_code=201,
)
async def upload_master_import(
    tenant_id: str,
    owner_module: str,
    master_key: str,
    file: Annotated[UploadFile, File()],
    effective_from: Annotated[date | None, Form(alias="effectiveFrom")] = None,
    idempotency_key: Annotated[str, Header(alias="Idempotency-Key", min_length=1)] = "",
    admin_request: Annotated[HumanAdminRequest, Depends(require_super_admin_request)] = None,  # type: ignore[assignment]
    connection: Annotated[Connection, Depends(get_connection)] = None,  # type: ignore[assignment]
) -> ProjectMasterImportResponse:
    set_tenant_context(connection, tenant_id)
    if owner_module != "AUDIT_CORE":
        raise AuditCoreError(
            error_code="VAC-MASTER-004",
            status_code=422,
            title="Owning module required",
            detail="DI-owned Project Master imports are supplied by DI through the facade.",
        )
    normalized = master_key.strip().upper()
    descriptor = descriptor_definition(normalized)
    if normalized not in excel_master_keys():
        raise AuditCoreError(
            error_code="VAC-MASTER-004",
            status_code=422,
            title="Excel administration unsupported",
            detail="The requested Project Master is not registered for Excel administration.",
        )
    if descriptor["requiresWef"] and effective_from is None:
        raise ValidationError(detail="effectiveFrom / WEF is required for this Project Master.")
    if effective_from is None:
        raise ValidationError(detail="effectiveFrom / WEF is required for this Project Master.")
    filename = file.filename or ""
    if not filename.lower().endswith(".xlsx"):
        raise ValidationError(detail="Project Master import must be an .xlsx workbook.")
    content = await file.read()
    if not content:
        raise ValidationError(detail="Project Master import workbook is empty.")
    file_hash = hashlib.sha256(content).hexdigest()
    template_version = _template_version(normalized)
    semantic_hash = stable_request_hash(
        {
            "ownerModule": owner_module,
            "masterKey": normalized,
            "effectiveFrom": effective_from.isoformat(),
            "templateVersion": template_version,
            "fileHash": file_hash,
        }
    )
    prior = connection.execute(
        text(
            """
            SELECT import_id, semantic_request_hash
            FROM auditcore.project_master_imports
            WHERE tenant_id=:tenant_id AND owner_module=:owner_module
              AND master_key=:master_key AND idempotency_key=:idempotency_key
            """
        ),
        {
            "tenant_id": tenant_id,
            "owner_module": owner_module,
            "master_key": normalized,
            "idempotency_key": idempotency_key,
        },
    ).mappings().one_or_none()
    if prior is not None:
        if prior["semantic_request_hash"] != semantic_hash:
            raise ConflictError(
                error_code="VAC-CONFLICT-003",
                title="Idempotency conflict",
                detail="Idempotency-Key was already used for a different master import.",
            )
        return _response(
            _import_row(connection, tenant_id=tenant_id, import_id=prior["import_id"])
        )

    import_id = connection.execute(
        text(
            """
            INSERT INTO auditcore.project_master_imports (
                tenant_id, owner_module, master_key, effective_from,
                template_version, original_file_name, file_hash,
                idempotency_key, semantic_request_hash, status,
                created_by_user_id
            ) VALUES (
                :tenant_id, :owner_module, :master_key, :effective_from,
                :template_version, :file_name, :file_hash,
                :idempotency_key, :semantic_hash, 'PARSING',
                :actor_id
            ) RETURNING import_id
            """
        ),
        {
            "tenant_id": tenant_id,
            "owner_module": owner_module,
            "master_key": normalized,
            "effective_from": effective_from,
            "template_version": template_version,
            "file_name": filename,
            "file_hash": file_hash,
            "idempotency_key": idempotency_key,
            "semantic_hash": semantic_hash,
            "actor_id": admin_request.user_id,
        },
    ).scalar_one()
    rows = _workbook_rows(content, master_key=normalized)
    validated = _validate_rows(
        connection,
        tenant_id=tenant_id,
        master_key=normalized,
        effective_from=effective_from,
        rows=rows,
    )
    valid_rows, warning_rows, error_rows = _insert_validated_rows(
        connection,
        tenant_id=tenant_id,
        import_id=import_id,
        rows=validated,
    )
    status = "VALIDATION_FAILED" if error_rows else "PREVIEW_READY"
    connection.execute(
        text(
            """
            UPDATE auditcore.project_master_imports
            SET status=:status, rows_parsed=:rows_parsed,
                valid_rows=:valid_rows, warning_rows=:warning_rows,
                error_rows=:error_rows, version_no=version_no+1
            WHERE tenant_id=:tenant_id AND import_id=:import_id
            """
        ),
        {
            "tenant_id": tenant_id,
            "import_id": import_id,
            "status": status,
            "rows_parsed": len(validated),
            "valid_rows": valid_rows,
            "warning_rows": warning_rows,
            "error_rows": error_rows,
        },
    )
    return _response(_import_row(connection, tenant_id=tenant_id, import_id=import_id))


@router.get(
    "/v1/tenants/{tenant_id}/project-master-imports/{import_id}",
    response_model=ProjectMasterImportResponse,
)
def get_master_import(
    tenant_id: str,
    import_id: UUID,
    admin_request: Annotated[HumanAdminRequest, Depends(require_super_admin_request)],
    connection: Annotated[Connection, Depends(get_connection)],
) -> ProjectMasterImportResponse:
    set_tenant_context(connection, tenant_id)
    return _response(_import_row(connection, tenant_id=tenant_id, import_id=import_id))


@router.get(
    "/v1/tenants/{tenant_id}/project-master-imports/{import_id}/rows",
    response_model=ImportRowsPage,
)
def get_master_import_rows(
    tenant_id: str,
    import_id: UUID,
    admin_request: Annotated[HumanAdminRequest, Depends(require_super_admin_request)],
    connection: Annotated[Connection, Depends(get_connection)],
    offset: Annotated[int, Query(ge=0)] = 0,
    limit: Annotated[int, Query(ge=1, le=200)] = 100,
    validation_status: Annotated[ValidationStatus | None, Query(alias="validationStatus")] = None,
) -> ImportRowsPage:
    set_tenant_context(connection, tenant_id)
    _import_row(connection, tenant_id=tenant_id, import_id=import_id)
    filters = ["tenant_id=:tenant_id", "import_id=:import_id"]
    params: dict[str, object] = {"tenant_id": tenant_id, "import_id": import_id}
    if validation_status is not None:
        filters.append("validation_status=:validation_status")
        params["validation_status"] = validation_status
    where = " AND ".join(filters)
    total = connection.execute(
        text(f"SELECT count(*) FROM auditcore.project_master_import_rows WHERE {where}"),
        params,
    ).scalar_one()
    rows = connection.execute(
        text(
            f"""
            SELECT row_number, parsed_data, validation_status, validation_messages
            FROM auditcore.project_master_import_rows
            WHERE {where}
            ORDER BY row_number
            OFFSET :offset LIMIT :limit
            """
        ),
        {**params, "offset": offset, "limit": limit},
    ).mappings().all()
    return ImportRowsPage(
        items=[
            ImportRowResponse(
                rowNumber=row["row_number"],
                parsedData=dict(row["parsed_data"]),
                validationStatus=row["validation_status"],
                messages=list(row["validation_messages"]),
            )
            for row in rows
        ],
        offset=offset,
        limit=limit,
        total=int(total),
    )


@router.get(
    "/v1/tenants/{tenant_id}/project-master-imports/{import_id}/error-report"
)
def download_import_error_report(
    tenant_id: str,
    import_id: UUID,
    admin_request: Annotated[HumanAdminRequest, Depends(require_super_admin_request)],
    connection: Annotated[Connection, Depends(get_connection)],
) -> StreamingResponse:
    set_tenant_context(connection, tenant_id)
    import_row = _import_row(connection, tenant_id=tenant_id, import_id=import_id)
    rows = connection.execute(
        text(
            """
            SELECT row_number, parsed_data, validation_status, validation_messages
            FROM auditcore.project_master_import_rows
            WHERE tenant_id=:tenant_id AND import_id=:import_id
              AND validation_status IN ('WARNING','ERROR')
            ORDER BY row_number
            """
        ),
        {"tenant_id": tenant_id, "import_id": import_id},
    ).mappings().all()
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Validation"
    columns = _TEMPLATE_COLUMNS[str(import_row["master_key"])]
    sheet.append(["row_number", "validation_status", "messages", *columns])
    for row in rows:
        parsed = dict(row["parsed_data"])
        sheet.append(
            [
                row["row_number"],
                row["validation_status"],
                " | ".join(row["validation_messages"]),
                *[parsed.get(column) for column in columns],
            ]
        )
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="import-{import_id}-validation.xlsx"'
        },
    )


@router.post(
    "/v1/tenants/{tenant_id}/project-master-imports/{import_id}/confirm",
    response_model=ProjectMasterImportResponse,
)
def confirm_master_import(
    tenant_id: str,
    import_id: UUID,
    admin_request: Annotated[HumanAdminRequest, Depends(require_super_admin_request)],
    connection: Annotated[Connection, Depends(get_connection)],
) -> ProjectMasterImportResponse:
    set_tenant_context(connection, tenant_id)
    current = _import_row(connection, tenant_id=tenant_id, import_id=import_id)
    if current["status"] == "CONFIRMED":
        return _response(current)
    if current["status"] not in {"PREVIEW_READY", "VALIDATION_FAILED"}:
        raise ConflictError(
            error_code="VAC-CONFLICT-008",
            title="Master import state conflict",
            detail="Project Master import is not ready for confirmation.",
        )
    if int(current["error_rows"]) > 0:
        raise AuditCoreError(
            error_code="VAC-MASTER-005",
            status_code=422,
            title="Master import has validation errors",
            detail="Correct blocking row errors before confirming this Project Master import.",
        )
    effective_from = current["effective_from"]
    if effective_from is None:
        raise AuditCoreError(
            error_code="VAC-MASTER-005",
            status_code=422,
            title="Master import has no WEF",
            detail="This effective-dated Project Master import requires an explicit WEF.",
        )
    rows = _load_parsed_rows(connection, tenant_id=tenant_id, import_id=import_id)
    version_id = _confirm_import(
        connection,
        tenant_id=tenant_id,
        master_key=current["master_key"],
        effective_from=effective_from,
        actor_id=admin_request.user_id,
        import_id=import_id,
        rows=rows,
    )
    connection.execute(
        text(
            """
            UPDATE auditcore.project_master_imports
            SET status='CONFIRMED', confirmed_version_id=:version_id,
                confirmed_by_user_id=:actor_id, confirmed_at_utc=now(),
                version_no=version_no+1
            WHERE tenant_id=:tenant_id AND import_id=:import_id
            """
        ),
        {
            "tenant_id": tenant_id,
            "import_id": import_id,
            "version_id": version_id,
            "actor_id": admin_request.user_id,
        },
    )
    return _response(_import_row(connection, tenant_id=tenant_id, import_id=import_id))


@router.delete(
    "/v1/tenants/{tenant_id}/project-master-imports/{import_id}",
    status_code=204,
)
def delete_master_import(
    tenant_id: str,
    import_id: UUID,
    admin_request: Annotated[HumanAdminRequest, Depends(require_super_admin_request)],
    connection: Annotated[Connection, Depends(get_connection)],
) -> None:
    set_tenant_context(connection, tenant_id)
    current = _import_row(connection, tenant_id=tenant_id, import_id=import_id)
    if current["status"] == "CONFIRMED":
        raise ConflictError(
            error_code="VAC-CONFLICT-008",
            title="Master import state conflict",
            detail="A confirmed import cannot be deleted through the staging delete operation.",
        )
    connection.execute(
        text(
            "DELETE FROM auditcore.project_master_imports "
            "WHERE tenant_id=:tenant_id AND import_id=:import_id"
        ),
        {"tenant_id": tenant_id, "import_id": import_id},
    )
